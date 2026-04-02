"""
VIP → Ohanafy Migration Workbook Populator

Connects to the Gulf PostgreSQL database and populates tabs in the
Distro Data Workbook from VIP source tables.

Usage:
    python3 populate_workbook.py [tab_name]

    tab_name: Optional. If provided, only populate that tab.
              e.g. "User", "Supplier", "Brands"
              If omitted, populates all implemented tabs.

Requires: .env file in the same directory with PostgreSQL credentials.
"""

import os
import sys
import psycopg2
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from copy import copy

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(SCRIPT_DIR, ".env")
WORKBOOK_PATH = os.path.join(SCRIPT_DIR, "Distro - Data Workbook TEMPLATE 3_3.xlsx")

# Styling constants (match existing workbook)
FILL_DATA = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
FILL_FORMULA = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FONT_DATA = Font(name="Calibri", size=11)

# ---------------------------------------------------------------------------
# Database connection
# ---------------------------------------------------------------------------
def get_connection():
    """Connect to PostgreSQL using .env credentials."""
    load_dotenv(ENV_PATH)
    return psycopg2.connect(
        host=os.getenv("PGHOST"),
        port=os.getenv("PGPORT"),
        dbname=os.getenv("PGDATABASE"),
        user=os.getenv("PGUSER"),
        password=os.getenv("PGPASSWORD"),
        sslmode=os.getenv("PGSSLMODE", "require"),
    )


def query(conn, sql, params=None):
    """Run a query and return rows as list of dicts."""
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def style_cell(cell, is_formula=False):
    """Apply standard data or formula styling to a cell."""
    cell.fill = FILL_FORMULA if is_formula else FILL_DATA
    cell.font = FONT_DATA
    cell.alignment = Alignment(vertical="center")


def clear_data_rows(ws, start_row=3):
    """Clear existing data rows (keep headers in rows 1-2)."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def s(val):
    """Safely strip a string value from the database."""
    if val is None:
        return None
    v = str(val).strip()
    return v if v else None


# ---------------------------------------------------------------------------
# Tab: User
# ---------------------------------------------------------------------------
def populate_user(conn, wb):
    """Populate the User tab from USERST + related tables."""
    ws = wb["User"]
    print("  Populating User tab...")

    # -- Query users --
    users = query(conn, """
        SELECT
            u."FIRSTNAME",
            u."LASTNAME",
            u."EMAIL",
            u."EMPLOYEEID",
            TRIM(u."EMPLOYEETYPE") AS employee_type,
            u."JOBTITLE",
            u."DEPARTMENTID",
            TRIM(u."DEFAULTWAREHOUSE") AS default_warehouse,
            TRIM(u."DEFAULTCOMPANY") AS default_company,
            u."EMPLOYEECODE",
            u."IBMIPROFILE",
            u."IDENTITY",
            u.import_is_deleted
        FROM analytics."USERST" u
        WHERE u.import_is_deleted = false
        ORDER BY u."LASTNAME", u."FIRSTNAME"
    """)

    # -- Query driver supervisors --
    drivers = query(conn, """
        SELECT TRIM("DRIVER") AS driver, TRIM("SUPERVISOR") AS supervisor
        FROM analytics."HDRDRT"
        WHERE import_is_deleted = false
    """)
    driver_supervisors = {d["driver"]: d["supervisor"] for d in drivers if d["supervisor"]}

    # -- Query merchandiser supervisors + phone --
    merchs = query(conn, """
        SELECT TRIM("MERCHANDISER") AS merch,
               TRIM("SUPERVISOR") AS supervisor,
               TRIM("PAGER"::text) AS pager
        FROM analytics."HDRMERCHT"
        WHERE import_is_deleted = false
    """)
    merch_supervisors = {m["merch"]: m["supervisor"] for m in merchs if m["supervisor"]}
    merch_phones = {m["merch"]: m["pager"] for m in merchs if m["pager"]}

    # -- Build employee code → full name lookup for supervisor resolution --
    emp_names = {}
    for u in users:
        code = s(u["EMPLOYEECODE"]) or s(u["IBMIPROFILE"])
        if code:
            fname = s(u["FIRSTNAME"]) or ""
            lname = s(u["LASTNAME"]) or ""
            emp_names[code.upper()] = f"{fname} {lname}".strip()

    # -- Role mapping from employee type --
    role_map = {
        "SLS": "Sales Rep",
        "DRIVR": "Driver",
        "MERCH": "Merchandiser",
        "SUPR": "Supervisor",
        "ADMIN": "Admin",
        "AREA": "Area Manager",
        "POS": "POS",
        "OTHER": "Other",
    }

    # -- Clear existing data --
    clear_data_rows(ws)

    print(f"  Writing {len(users)} users...")
    for i, u in enumerate(users):
        row = i + 3  # data starts at row 3
        emp_code = s(u["EMPLOYEECODE"]) or s(u["IBMIPROFILE"]) or ""
        emp_type = u["employee_type"] or ""

        # Look up supervisor
        supervisor_code = driver_supervisors.get(emp_code.upper()) or merch_supervisors.get(emp_code.upper())
        supervisor_name = emp_names.get(supervisor_code.upper()) if supervisor_code else None

        # Phone from merchandiser table
        phone = merch_phones.get(emp_code.upper())

        # A: First Name
        ws.cell(row=row, column=1, value=s(u["FIRSTNAME"]))
        style_cell(ws.cell(row=row, column=1))

        # B: Last Name
        ws.cell(row=row, column=2, value=s(u["LASTNAME"]))
        style_cell(ws.cell(row=row, column=2))

        # C: Email
        ws.cell(row=row, column=3, value=s(u["EMAIL"]))
        style_cell(ws.cell(row=row, column=3))

        # D: Role
        ws.cell(row=row, column=4, value=role_map.get(emp_type, ""))
        style_cell(ws.cell(row=row, column=4))

        # E: Phone
        ws.cell(row=row, column=5, value=phone)
        style_cell(ws.cell(row=row, column=5))

        # F: Employee ID
        ws.cell(row=row, column=6, value=s(u["EMPLOYEEID"]))
        style_cell(ws.cell(row=row, column=6))

        # G: Employee Type
        ws.cell(row=row, column=7, value=emp_type)
        style_cell(ws.cell(row=row, column=7))

        # H: Job Title
        ws.cell(row=row, column=8, value=s(u["JOBTITLE"]))
        style_cell(ws.cell(row=row, column=8))

        # I: Department
        ws.cell(row=row, column=9, value=s(u["DEPARTMENTID"]))
        style_cell(ws.cell(row=row, column=9))

        # J: Default Warehouse
        ws.cell(row=row, column=10, value=s(u["default_warehouse"]))
        style_cell(ws.cell(row=row, column=10))

        # K: Default Company
        ws.cell(row=row, column=11, value=s(u["default_company"]))
        style_cell(ws.cell(row=row, column=11))

        # L: Supervisor
        ws.cell(row=row, column=12, value=supervisor_name)
        style_cell(ws.cell(row=row, column=12))

        # M: Active
        ws.cell(row=row, column=13, value="TRUE")
        style_cell(ws.cell(row=row, column=13))

        # N: VIP User Code
        ws.cell(row=row, column=14, value=emp_code)
        style_cell(ws.cell(row=row, column=14))

        # O: Alias (formula)
        ws.cell(row=row, column=15, value=f'=LEFT(A{row},1)&B{row}')
        style_cell(ws.cell(row=row, column=15), is_formula=True)

        # P: Username (formula = email)
        ws.cell(row=row, column=16, value=f'=C{row}')
        style_cell(ws.cell(row=row, column=16), is_formula=True)

        # Q: Time Zone
        ws.cell(row=row, column=17, value="America/Chicago")
        style_cell(ws.cell(row=row, column=17), is_formula=True)

        # R: Profile Name
        profile = "System Administrator" if emp_type == "ADMIN" else "Ohanafy Standard User"
        ws.cell(row=row, column=18, value=profile)
        style_cell(ws.cell(row=row, column=18), is_formula=True)

        # S: User Name (formula)
        ws.cell(row=row, column=19, value=f'=A{row}&" "&B{row}')
        style_cell(ws.cell(row=row, column=19))

        # T: Mapping Key (formula)
        ws.cell(row=row, column=20, value=f'=CONCAT("U-",Y{row})')
        style_cell(ws.cell(row=row, column=20))

        # U: Profile Id (VLOOKUP against Profiles tab)
        ws.cell(row=row, column=21, value=f"=VLOOKUP(R{row},Profiles!$A$2:$B$3,2,FALSE)")
        style_cell(ws.cell(row=row, column=21))

        # V: LocaleSidKey
        ws.cell(row=row, column=22, value="en_US")
        style_cell(ws.cell(row=row, column=22))

        # W: LanguageLocaleKey
        ws.cell(row=row, column=23, value="en_US")
        style_cell(ws.cell(row=row, column=23))

        # X: EmailEncodingKey
        ws.cell(row=row, column=24, value="UTF-8")
        style_cell(ws.cell(row=row, column=24))

        # Y: Record Number
        ws.cell(row=row, column=25, value=i + 1)
        style_cell(ws.cell(row=row, column=25))

        # Z: True
        ws.cell(row=row, column=26, value=True)
        style_cell(ws.cell(row=row, column=26))

    print(f"  ✓ User tab: {len(users)} rows written")
    return len(users)


# ---------------------------------------------------------------------------
# Helpers: Column insertion
# ---------------------------------------------------------------------------
FILL_HEADER = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
FILL_API = PatternFill(start_color="FF5F9EA0", end_color="FF5F9EA0", fill_type="solid")
FILL_SYSTEM = PatternFill(start_color="FF3F5372", end_color="FF3F5372", fill_type="solid")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
FONT_API = Font(name="Calibri", size=11, color="FFFFFFFF")


def insert_columns_with_headers(ws, insert_at, columns):
    """Insert new columns at position and set row 1/2 headers.

    columns: list of (header_name, api_name_or_None)
        If api_name is None, uses FILL_SYSTEM styling for row 2.
    """
    num_cols = len(columns)
    ws.insert_cols(insert_at, num_cols)

    for i, (header, api_name) in enumerate(columns):
        col = insert_at + i

        # Row 1: header
        cell1 = ws.cell(row=1, column=col, value=header)
        cell1.fill = FILL_HEADER
        cell1.font = FONT_HEADER
        cell1.alignment = Alignment(vertical="center")

        # Row 2: API name
        cell2 = ws.cell(row=2, column=col, value=api_name)
        cell2.fill = FILL_SYSTEM if api_name is None else FILL_API
        cell2.font = FONT_API
        cell2.alignment = Alignment(vertical="center")


# ---------------------------------------------------------------------------
# Tab: ParentLocation
# ---------------------------------------------------------------------------
def populate_parent_location(conn, wb):
    """Populate the ParentLocation tab from HDRWHSET + HDRCOMPT."""
    ws = wb["ParentLocation"]
    print("  Populating ParentLocation tab...")

    # -- Insert 3 new columns at H (col 8), pushing Record Number to K --
    # Check if already inserted by looking at current H header
    if ws.cell(row=1, column=8).value != "VIP Warehouse Code":
        insert_columns_with_headers(ws, 8, [
            ("VIP Warehouse Code", "ohanafy__VIP_Warehouse_Code__c"),
            ("Company Name", "ohanafy__Company_Name__c"),
            ("GL Company", "ohanafy__GL_Company__c"),
        ])
        # Fix Mapping Key formula in B: was =CONCAT("PL-",H3), now Record Number is at K
        # (will be set per-row below)
        print("  Inserted 3 new columns (H-J): VIP Warehouse Code, Company Name, GL Company")

    # -- Query warehouses --
    warehouses = query(conn, """
        SELECT TRIM("WAREHOUSE") AS wh_code,
               TRIM("WAREHOUSE_NAME") AS wh_name,
               TRIM("GL_COMPANY") AS gl_company,
               TRIM("DELETE_FLAG") AS del_flag
        FROM analytics."HDRWHSET"
        WHERE import_is_deleted = false
        ORDER BY "WAREHOUSE"
    """)

    # -- Query companies for name lookup --
    companies = query(conn, """
        SELECT TRIM("COMPANY") AS co_code,
               TRIM("COMPANY_NAME") AS co_name
        FROM analytics."HDRCOMPT"
        WHERE import_is_deleted = false
    """)
    co_names = {c["co_code"]: c["co_name"] for c in companies}

    # -- Clear existing data --
    clear_data_rows(ws)

    # -- Filter: exclude deleted warehouses (DELETE_FLAG = 'X') --
    active_wh = [w for w in warehouses if w["del_flag"] != "X"]

    print(f"  Writing {len(active_wh)} parent locations (of {len(warehouses)} total, {len(warehouses)-len(active_wh)} deleted)...")
    for i, w in enumerate(active_wh):
        row = i + 3

        # A: Location Name
        ws.cell(row=row, column=1, value=w["wh_name"])
        style_cell(ws.cell(row=row, column=1))

        # B: Mapping Key (formula, Record Number now at K)
        ws.cell(row=row, column=2, value=f'=CONCAT("PL-",K{row})')
        style_cell(ws.cell(row=row, column=2))

        # C: Finished Good
        ws.cell(row=row, column=3, value=True)
        style_cell(ws.cell(row=row, column=3))

        # D: Keg Shell
        ws.cell(row=row, column=4, value=True)
        style_cell(ws.cell(row=row, column=4))

        # E: Merchandise
        ws.cell(row=row, column=5, value=True)
        style_cell(ws.cell(row=row, column=5))

        # F: Tap Handle
        ws.cell(row=row, column=6, value=True)
        style_cell(ws.cell(row=row, column=6))

        # G: Active
        ws.cell(row=row, column=7, value=True)
        style_cell(ws.cell(row=row, column=7))

        # H: VIP Warehouse Code (NEW)
        ws.cell(row=row, column=8, value=w["wh_code"])
        style_cell(ws.cell(row=row, column=8))

        # I: Company Name (NEW)
        # Map warehouse's GL_COMPANY to company name
        co_name = co_names.get(w["wh_code"]) or co_names.get(w["gl_company"])
        ws.cell(row=row, column=9, value=co_name)
        style_cell(ws.cell(row=row, column=9))

        # J: GL Company (NEW)
        ws.cell(row=row, column=10, value=w["gl_company"])
        style_cell(ws.cell(row=row, column=10))

        # K: Record Number
        ws.cell(row=row, column=11, value=i + 1)
        style_cell(ws.cell(row=row, column=11))

    print(f"  ✓ ParentLocation tab: {len(active_wh)} rows written")
    return len(active_wh)


# ---------------------------------------------------------------------------
# Tab: ChildLocation
# ---------------------------------------------------------------------------
def populate_child_location(conn, wb):
    """Populate the ChildLocation tab from LOCMAST zones + HDRWHSET."""
    ws = wb["ChildLocation"]
    print("  Populating ChildLocation tab...")

    # -- Insert new columns before Record Number (col L=12) --
    # Add: VIP Warehouse Code, Zone Description, Bin Count
    if ws.cell(row=1, column=12).value != "VIP Warehouse Code":
        insert_columns_with_headers(ws, 12, [
            ("VIP Warehouse Code", "ohanafy__VIP_Warehouse_Code__c"),
            ("Zone Description", None),
            ("Bin Count", None),
        ])
        print("  Inserted 3 new columns (L-N): VIP Warehouse Code, Zone Description, Bin Count")

    # Record Number is now at col O (15), True at col P (16)
    # Site Key VLOOKUP in K also needs updating for new ParentLocation range

    # -- Get warehouse names for parent location mapping --
    warehouses = query(conn, """
        SELECT TRIM("WAREHOUSE") AS wh_code,
               TRIM("WAREHOUSE_NAME") AS wh_name,
               TRIM("DELETE_FLAG") AS del_flag
        FROM analytics."HDRWHSET"
        WHERE import_is_deleted = false
        ORDER BY "WAREHOUSE"
    """)
    wh_names = {w["wh_code"]: w["wh_name"] for w in warehouses if w["del_flag"] != "X"}

    # -- Query distinct zones per warehouse --
    # Group LOCMAST by warehouse + description to get logical zones
    # Exclude WH 4 (has product descriptions instead of zone names)
    # Require at least 1 bin per zone
    zones = query(conn, """
        SELECT TRIM("LCWHSE") AS wh,
               TRIM("LCDESC") AS zone_desc,
               count(*) AS bin_count,
               bool_or("LCPALL"::text = 'Y') AS has_pallets,
               bool_or("LCSTAG"::text = 'Y') AS has_staging,
               bool_or("LCBOND"::text = 'Y') AS has_bond
        FROM analytics."LOCMAST"
        WHERE import_is_deleted = false
          AND TRIM("LCSTAT") = 'Y'
          AND TRIM("LCDESC") <> ''
          AND TRIM("LCWHSE") <> '4'
          AND TRIM("LCWHSE") <> 'c'
        GROUP BY TRIM("LCWHSE"), TRIM("LCDESC")
        HAVING count(*) >= 2
           OR TRIM("LCDESC") IN (
               'POINT OF SALE', 'RECEIVING DOCK', 'DUNNAGE',
               'DEPOT STAGING', 'LIQUOR', 'POS'
           )
        ORDER BY TRIM("LCWHSE"), count(*) DESC
    """)

    # -- Build child location names --
    # Normalize zone names and determine product flags
    def is_keg_zone(desc):
        d = desc.upper()
        return "KEG" in d or "COOLER" in d or "COLD" in d

    def is_merch_zone(desc):
        d = desc.upper()
        return "POS" in d or "MERCHANDISE" in d or "POINT OF SALE" in d

    # -- Clear existing data --
    clear_data_rows(ws)

    row_num = 0
    for z in zones:
        wh_code = z["wh"]
        wh_name = wh_names.get(wh_code)
        if not wh_name:
            continue  # skip warehouses not in ParentLocation

        row_num += 1
        row = row_num + 2  # data starts at row 3
        zone_name = z["zone_desc"]
        child_name = f"{zone_name}"

        keg = is_keg_zone(zone_name)
        merch = is_merch_zone(zone_name)

        # A: Record Type Name
        ws.cell(row=row, column=1, value="Physical")
        style_cell(ws.cell(row=row, column=1))

        # B: RecordTypeId (VLOOKUP)
        ws.cell(row=row, column=2, value=f"=VLOOKUP(A{row},'Record Types'!$I$3:$J$4,2,FALSE)")
        style_cell(ws.cell(row=row, column=2))

        # C: Site Name (Parent Location) — must match ParentLocation col A
        ws.cell(row=row, column=3, value=wh_name)
        style_cell(ws.cell(row=row, column=3))

        # D: Child Location Name
        ws.cell(row=row, column=4, value=child_name)
        style_cell(ws.cell(row=row, column=4))

        # E: Mapping Key (formula — Record Number now at O)
        ws.cell(row=row, column=5, value=f'=CONCAT("CL-",O{row})')
        style_cell(ws.cell(row=row, column=5))

        # F: Finished Good
        ws.cell(row=row, column=6, value=not merch)
        style_cell(ws.cell(row=row, column=6))

        # G: Merchandise
        ws.cell(row=row, column=7, value=merch)
        style_cell(ws.cell(row=row, column=7))

        # H: Keg Shell
        ws.cell(row=row, column=8, value=keg)
        style_cell(ws.cell(row=row, column=8))

        # I: Tap Handle
        ws.cell(row=row, column=9, value=False)
        style_cell(ws.cell(row=row, column=9))

        # J: Active
        ws.cell(row=row, column=10, value=True)
        style_cell(ws.cell(row=row, column=10))

        # K: Site Key (VLOOKUP to ParentLocation — updated range for 9 parent rows + 3 new cols)
        ws.cell(row=row, column=11,
                value=f'=IF(ISNA(VLOOKUP(C{row},ParentLocation!$A$3:$B$11,2,FALSE))=TRUE,"",VLOOKUP(C{row},ParentLocation!$A$3:$B$11,2,FALSE))')
        style_cell(ws.cell(row=row, column=11))

        # L: VIP Warehouse Code (NEW)
        ws.cell(row=row, column=12, value=wh_code)
        style_cell(ws.cell(row=row, column=12))

        # M: Zone Description (NEW)
        ws.cell(row=row, column=13, value=zone_name)
        style_cell(ws.cell(row=row, column=13))

        # N: Bin Count (NEW)
        ws.cell(row=row, column=14, value=z["bin_count"])
        style_cell(ws.cell(row=row, column=14))

        # O: Record Number
        ws.cell(row=row, column=15, value=row_num)
        style_cell(ws.cell(row=row, column=15))

        # P: True
        ws.cell(row=row, column=16, value=True)
        style_cell(ws.cell(row=row, column=16))

    print(f"  ✓ ChildLocation tab: {row_num} rows written")
    return row_num


# ---------------------------------------------------------------------------
# Tab: Product
# ---------------------------------------------------------------------------
def populate_product(conn, wb):
    """Populate the Product tab from ITEMT + BRANDT + SUPPLIERT."""
    ws = wb["Product"]
    print("  Populating Product tab...")

    # -- Insert 6 new columns at AG (col 33), pushing Record Number to AM (39) --
    if ws.cell(row=1, column=33).value != "VIP Item Code":
        insert_columns_with_headers(ws, 33, [
            ("VIP Item Code", "ohanafy__VIP_Item_Code__c"),
            ("Case UPC", "ohanafy__Case_UPC__c"),
            ("VIP Supplier Code", None),
            ("VIP Brand Code", None),
            ("Container Type", None),
            ("Product Class", None),
        ])
        print("  Inserted 6 new columns (AG-AL)")

    # Record Number is now col AM (39)
    REC_NUM_COL = 39

    # -- Build brand lookup: brand_code → (brand_name, supplier_code) --
    brands = query(conn, """
        SELECT TRIM("BRAND") AS code, TRIM("BRAND_NAME") AS name,
               TRIM("SUPPLIER") AS supplier
        FROM analytics."BRANDT"
        WHERE import_is_deleted = false
    """)
    brand_map = {b["code"]: b for b in brands}

    # -- Build supplier lookup: supplier_code → supplier_name --
    suppliers = query(conn, """
        SELECT TRIM("SUPPLIER") AS code, TRIM("SUPPLIER_NAME") AS name
        FROM analytics."SUPPLIERT"
    """)
    supplier_map = {s["code"]: s["name"] for s in suppliers}

    # -- Build item type lookup: product_type_pointer → description --
    item_types = query(conn, """
        SELECT "IDENTITY"::text AS id, TRIM("DESCRIPTION") AS desc,
               TRIM("INVOICETYPE") AS inv_type
        FROM analytics."HDRITYPET"
        WHERE import_is_deleted = false
    """)
    itype_map = {t["id"]: t for t in item_types}

    # -- Query all active + seasonal items --
    items = query(conn, """
        SELECT TRIM("ITEM_CODE") AS item_code,
               TRIM("PACKAGE_DESCRIPTION") AS pkg_desc,
               TRIM("BRAND_CODE") AS brand_code,
               TRIM("PRODUCT_CLASS"::text) AS product_class,
               TRIM("ITEM_STATUS") AS status,
               TRIM("CONTAINER_TYPE_CODE") AS container_type,
               TRIM("PACKAGE_SIZE") AS pkg_size,
               "UNITS_CASE" AS units_case,
               "ALCOHOL_PERCENT" AS abv,
               TRIM("SUPPLIER_CODE") AS supplier_code,
               TRIM("TAXABLE_FLAG") AS taxable,
               "SUGGESTED_SELLING_PRICE" AS ssp,
               "PRODUCT_TYPE_POINTER"::text AS prod_type_ptr,
               TRIM("PACKAGE_SHORT") AS pkg_short,
               TRIM("STYLE_TYPE_CODE") AS style_code,
               "CASE_UPC"::text AS case_upc,
               TRIM("PRODUCT_CODE"::text) AS prod_code,
               "MILLILITERS_PER_BOTTLE" AS ml_per_bottle,
               TRIM("SELLABLE_BY_UNIT") AS sell_by_unit
        FROM analytics."ITEMT"
        WHERE import_is_deleted = false
          AND TRIM("ITEM_STATUS") IN ('A', 'S')
        ORDER BY "BRAND_CODE", "ITEM_CODE"
    """)

    # -- Determine Ohanafy type from VIP data --
    def get_ohanafy_type(item):
        desc = (item["pkg_desc"] or "").upper()
        ct = item["container_type"] or ""
        ptp = item["prod_type_ptr"] or ""
        pclass = item["product_class"] or ""

        if "DRAFT SUPPLIES" in desc or "TAP HANDLE" in desc:
            return "Tap Handle"
        if "COOPERAGE" in desc or "EMPTY KEG" in desc or "SHELL" in desc:
            return "Keg Shell"
        if pclass == "40" or ptp == "11":
            return "Merchandise"
        return "Finished Good"

    def get_sub_type(item):
        ct = item["container_type"] or ""
        if ct in ("710", "720", "730"):  # BBL sizes
            return "Draft"
        return "Packaged"

    def get_item_purpose(item):
        status = item["status"] or ""
        if status == "S":
            return "Buy/Sell"
        return "Buy/Sell"

    # -- Clear existing data --
    clear_data_rows(ws)

    # Item Line and Brands tab ranges for VLOOKUPs
    il_max = 929  # Suppliers(Item Line) has 927 data rows starting at row 3
    br_max = 928  # Brands has 926 data rows starting at row 3
    cl_max = 52   # ChildLocation has 50 data rows starting at row 3

    print(f"  Writing {len(items)} products...")
    for i, item in enumerate(items):
        row = i + 3
        brand_info = brand_map.get(item["brand_code"], {})
        brand_name = brand_info.get("name", "")
        supp_code = brand_info.get("supplier", item["supplier_code"] or "")
        ohanafy_type = get_ohanafy_type(item)
        sub_type = get_sub_type(item)
        units = item["units_case"] or 0

        # A: Record Type Name (formula = same as Type in col N)
        ws.cell(row=row, column=1, value=f"=N{row}")
        style_cell(ws.cell(row=row, column=1))

        # B: Record Type Id (VLOOKUP to Record Types)
        ws.cell(row=row, column=2,
                value=f"=VLOOKUP(A{row},'Record Types'!$G$3:$H$8,2,FALSE)")
        style_cell(ws.cell(row=row, column=2))

        # C: Product Name
        ws.cell(row=row, column=3, value=s(item["pkg_desc"]))
        style_cell(ws.cell(row=row, column=3))

        # D: Mapping Key (Record Number now at col AM=39)
        ws.cell(row=row, column=4, value=f'=CONCAT("FG-",AM{row})')
        style_cell(ws.cell(row=row, column=4))

        # E: Supplier (Item Line Name) — brand name (1:1 mapping)
        ws.cell(row=row, column=5, value=brand_name)
        style_cell(ws.cell(row=row, column=5))

        # F: Brand — brand name
        ws.cell(row=row, column=6, value=brand_name)
        style_cell(ws.cell(row=row, column=6))

        # G: Units
        ws.cell(row=row, column=7, value=units if units else None)
        style_cell(ws.cell(row=row, column=7))

        # H: Units Per Case
        ws.cell(row=row, column=8, value=units if units else None)
        style_cell(ws.cell(row=row, column=8))

        # I: Default Location Name (leave blank — needs warehouse-level config)
        ws.cell(row=row, column=9, value=None)
        style_cell(ws.cell(row=row, column=9))

        # J: Default Price (SSP)
        ssp = item["ssp"] or 0
        ws.cell(row=row, column=10, value=float(ssp) if ssp else None)
        style_cell(ws.cell(row=row, column=10))

        # K: Last Sale Price (leave blank)
        ws.cell(row=row, column=11, value=None)
        style_cell(ws.cell(row=row, column=11))

        # L: Style
        ws.cell(row=row, column=12, value=s(item["style_code"]))
        style_cell(ws.cell(row=row, column=12))

        # M: Category (from product class — will be descriptive)
        ws.cell(row=row, column=13, value=s(item["product_class"]))
        style_cell(ws.cell(row=row, column=13))

        # N: Type
        ws.cell(row=row, column=14, value=ohanafy_type)
        style_cell(ws.cell(row=row, column=14))

        # O: Item Purpose
        ws.cell(row=row, column=15, value=get_item_purpose(item))
        style_cell(ws.cell(row=row, column=15))

        # P: Packaging Style
        ws.cell(row=row, column=16, value=s(item["pkg_short"]))
        style_cell(ws.cell(row=row, column=16))

        # Q: UOM Sub Type (package size description)
        ws.cell(row=row, column=17, value=s(item["pkg_size"]))
        style_cell(ws.cell(row=row, column=17))

        # R: ABV
        abv = item["abv"] or 0
        ws.cell(row=row, column=18, value=float(abv) if abv > 0 else None)
        style_cell(ws.cell(row=row, column=18))

        # S: AVG COGS (leave blank — need cost table)
        ws.cell(row=row, column=19, value=None)
        style_cell(ws.cell(row=row, column=19))

        # T: Item Purpose Key (formula)
        ws.cell(row=row, column=20,
                value=f'=IF(O{row}="Buy","Buy",IF(O{row}="Sell","Sell",IF(O{row}="Buy/Sell","Buy;Sell","")))')
        style_cell(ws.cell(row=row, column=20), is_formula=True)

        # U: UOM (formula based on Q — simplified from template)
        ws.cell(row=row, column=21,
                value=f'=IF(OR(ISNUMBER(SEARCH("BBL",Q{row})),ISNUMBER(SEARCH("Liter",Q{row})),ISNUMBER(SEARCH("GAL",Q{row}))),"US Volume",IF(ISNUMBER(SEARCH("ML",Q{row})),"Metric Volume",IF(ISNUMBER(SEARCH("Each",Q{row})),"US Count","US Volume")))')
        style_cell(ws.cell(row=row, column=21), is_formula=True)

        # V: Tax Exempt (inverse of taxable flag)
        taxable = (item["taxable"] or "").upper()
        ws.cell(row=row, column=22, value=taxable != "Y")
        style_cell(ws.cell(row=row, column=22))

        # W: Active
        ws.cell(row=row, column=23, value=True)
        style_cell(ws.cell(row=row, column=23))

        # X: Sub Type
        ws.cell(row=row, column=24, value=sub_type)
        style_cell(ws.cell(row=row, column=24))

        # Y: Avg Production Cost Total (=S)
        ws.cell(row=row, column=25, value=f"=S{row}")
        style_cell(ws.cell(row=row, column=25), is_formula=True)

        # Z: Avg Production Cost (=S)
        ws.cell(row=row, column=26, value=f"=S{row}")
        style_cell(ws.cell(row=row, column=26), is_formula=True)

        # AA: Avg Production Cost Counter
        ws.cell(row=row, column=27, value=1)
        style_cell(ws.cell(row=row, column=27))

        # AB: Shelf Life Days
        ws.cell(row=row, column=28, value=None)
        style_cell(ws.cell(row=row, column=28))

        # AC: State Product Code
        ws.cell(row=row, column=29, value=s(item["prod_code"]))
        style_cell(ws.cell(row=row, column=29))

        # AD: Item Line Key (VLOOKUP to Suppliers(Item Line) tab)
        ws.cell(row=row, column=30,
                value=f"=IF(ISNA(VLOOKUP(E{row},'Suppliers(Item Line)'!$A$3:$H${il_max},8,FALSE)),\"\",VLOOKUP(E{row},'Suppliers(Item Line)'!$A$3:$H${il_max},8,FALSE))")
        style_cell(ws.cell(row=row, column=30), is_formula=True)

        # AE: Item Type Key (VLOOKUP to Brands tab — col C=Name, col D=MappingKey)
        ws.cell(row=row, column=31,
                value=f"=IF(ISNA(VLOOKUP(F{row},Brands!$C$3:$D${br_max},2,FALSE)),\"\",VLOOKUP(F{row},Brands!$C$3:$D${br_max},2,FALSE))")
        style_cell(ws.cell(row=row, column=31), is_formula=True)

        # AF: Default Location Key (VLOOKUP to ChildLocation)
        ws.cell(row=row, column=32,
                value=f"=IF(ISNA(VLOOKUP(I{row},ChildLocation!$D$3:$E${cl_max},2,FALSE)),\"\",VLOOKUP(I{row},ChildLocation!$D$3:$E${cl_max},2,FALSE))")
        style_cell(ws.cell(row=row, column=32), is_formula=True)

        # AG-AL: New columns (33-38)
        # AG: VIP Item Code
        ws.cell(row=row, column=33, value=s(item["item_code"]))
        style_cell(ws.cell(row=row, column=33))

        # AH: Case UPC
        upc = item["case_upc"] or "0"
        ws.cell(row=row, column=34, value=upc if upc != "0" else None)
        style_cell(ws.cell(row=row, column=34))

        # AI: VIP Supplier Code
        ws.cell(row=row, column=35, value=supp_code)
        style_cell(ws.cell(row=row, column=35))

        # AJ: VIP Brand Code
        ws.cell(row=row, column=36, value=s(item["brand_code"]))
        style_cell(ws.cell(row=row, column=36))

        # AK: Container Type
        ws.cell(row=row, column=37, value=s(item["container_type"]))
        style_cell(ws.cell(row=row, column=37))

        # AL: Product Class
        ws.cell(row=row, column=38, value=s(item["product_class"]))
        style_cell(ws.cell(row=row, column=38))

        # AM: Record Number (col 39)
        ws.cell(row=row, column=39, value=i + 1)
        style_cell(ws.cell(row=row, column=39))

    # Type distribution summary
    type_counts = {}
    for item in items:
        t = get_ohanafy_type(item)
        type_counts[t] = type_counts.get(t, 0) + 1
    print(f"  Type breakdown: {type_counts}")
    print(f"  ✓ Product tab: {len(items)} rows written")
    return len(items)


# ---------------------------------------------------------------------------
# Tab: Customer
# ---------------------------------------------------------------------------
def populate_customer(conn, wb):
    """Populate the Customer tab from BRATTT + POSACCT + CITYT + lookups."""
    ws = wb["Customer"]
    print("  Populating Customer tab...")

    # -- Insert 6 new VIP reference columns before Record Number (col 38=AL) --
    if ws.cell(row=1, column=38).value != "VIP On/Off Code":
        insert_columns_with_headers(ws, 38, [
            ("VIP On/Off Code", None),
            ("VIP Route Code", None),
            ("VIP Chain Code", None),
            ("VIP Class Code", None),
            ("VIP Salesman Code", None),
            ("VIP County Code", None),
        ])
        print("  Inserted 6 new columns (AL-AQ): VIP reference fields")

    # Record Number is now col AR (44)
    REC_NUM_COL = 44

    # -- Build lookup tables --

    # City lookup: ID_CITY → (city_name, state)
    cities = query(conn, """
        SELECT "IDENTITY" AS id, TRIM("CITYNAME") AS city, TRIM("STATE") AS state
        FROM analytics."CITYT" WHERE import_is_deleted = false
    """)
    city_map = {c["id"]: c for c in cities}

    # Payment terms: CMTERM code → standard Ohanafy term
    # CMTERM stores TERMSCODE (E,N,5,H,0) or VIP-internal codes (1,2,3,F,8,9,6)
    terms_map = {
        "0": "COD",          # MONEY ORDER/CHECK, 0 days
        "1": "Net 30",       # Standard charge (HDRTERMT ID=1, CHARGE EDI)
        "2": "Net 15",       # 15-day charge terms
        "3": "COD",          # Check on delivery
        "5": "COD",          # BEER COD N/A CHARGE
        "6": "Net 30",       # Special charge terms
        "8": "COD",          # BEER COD (HDRTERMT ID=8)
        "9": "COD",          # Credit hold / special COD
        "E": "Net 30",       # EDI charge, 30 days
        "F": "Net 30",       # FinTech/EFT charge
        "H": "Net 60",       # NET 60/1% DISCOUNT
        "N": "Net 15",       # NET 15 DAYS CHARGE
    }

    # On/Off premise lookup (simplified to On/Off only)
    onoff_map = {"1": "On Premise", "2": "Off Premise", "3": "Off Premise"}

    # Payment method: CMETHO code → method name, plus override from term code
    method_map = {
        "4": "Check",
        "W": "Wire",
        "2": "Cash",
        "M": "Money Order",
        "A": "ACH",
        "O": "Other",
    }
    # Term-based payment method overrides (EDI/FinTech customers)
    term_method_map = {
        "E": "EDI",
        "F": "FinTech",
    }

    # County lookup: county code → county name
    counties = query(conn, """
        SELECT "COUNTY"::text AS code, TRIM("COUNTY_NAME") AS name
        FROM analytics."HDRCNTYT" WHERE import_is_deleted = false
    """)
    county_map = {c["code"].strip(): c["name"] for c in counties}

    # Chain lookup: chain code → description
    chains = query(conn, """
        SELECT "CHAIN"::text AS code, TRIM("DESCRIPTION") AS desc
        FROM analytics."HDRCHAINT" WHERE import_is_deleted = false
    """)
    chain_map = {c["code"].strip(): c["desc"] for c in chains}

    # -- Salesman code → full name (best available source) --
    # Priority: USERST SLS (real names) → POSSM → HDRDRT
    placeholder_words = {"OPEN", "DRIVER SELL", "HOUSE", "WEBORDERS", "MOBILE HOUSE",
                         "MILTON HOUSE", "BHAM HOUSE", "MONTG HOUSE"}

    def is_placeholder(name):
        upper = (name or "").upper().strip()
        return not upper or any(upper.startswith(w) or upper == w for w in placeholder_words)

    def flip_last_first(name):
        """Convert 'LAST, FIRST' to 'First Last' title case."""
        parts = name.split(",", 1)
        if len(parts) == 2:
            return f"{parts[1].strip()} {parts[0].strip()}".title()
        return name.strip().title()

    # 1. USERST SLS type with real names
    sls_reps = query(conn, """
        SELECT TRIM("EMPLOYEECODE") AS code,
               TRIM("FIRSTNAME") || ' ' || TRIM("LASTNAME") AS name
        FROM analytics."USERST"
        WHERE import_is_deleted = false AND TRIM("EMPLOYEETYPE") = 'SLS'
          AND TRIM("EMPLOYEECODE") <> ''
    """)
    rep_map = {}
    for r in sls_reps:
        if not is_placeholder(r["name"]):
            rep_map[r["code"].upper()] = r["name"]

    # 2. POSSM (POS Salesman) — "LAST, FIRST" format
    possm_reps = query(conn, """
        SELECT TRIM("SALESMAN_CODE") AS code, TRIM("SALESMAN_NAME") AS name
        FROM analytics."POSSM" WHERE import_is_deleted = false
    """)
    for r in possm_reps:
        code = r["code"].upper()
        if code not in rep_map and not is_placeholder(r["name"]):
            rep_map[code] = flip_last_first(r["name"])

    # 3. HDRDRT (drivers) — "LAST, FIRST" format
    drv_reps = query(conn, """
        SELECT TRIM("DRIVER") AS code, TRIM("DRIVER_NAME") AS name
        FROM analytics."HDRDRT" WHERE import_is_deleted = false
    """)
    for r in drv_reps:
        code = r["code"].upper()
        if code not in rep_map and not is_placeholder(r["name"]):
            rep_map[code] = flip_last_first(r["name"])

    # Pricelist price code → pricelist name (from workbook Pricelist tab)
    ws_pl = wb["Pricelist"]
    price_code_to_name = {}
    for row in range(3, ws_pl.max_row + 1):
        name = ws_pl.cell(row=row, column=1).value   # col A: Name
        pcode = ws_pl.cell(row=row, column=7).value   # col G: VIP Price Code
        if pcode and name:
            price_code_to_name[str(pcode).strip()] = str(name).strip()

    # -- Addresses: POSACCT (delivery) + CUSTADT (mailing fallback) --
    addrs = query(conn, """
        SELECT TRIM("CUSTOMER_ID") AS acct,
               TRIM("DEL_ADDRESS1") AS addr1,
               TRIM("DEL_ADDRESS2") AS addr2,
               TRIM("DEL_CITY") AS city,
               TRIM("DEL_STATE") AS state,
               TRIM("DEL_ZIP") AS zip
        FROM analytics."POSACCT"
        WHERE import_is_deleted = false
    """)
    addr_map = {a["acct"]: a for a in addrs}

    # CUSTADT mailing addresses as fallback
    mail_addrs = query(conn, """
        SELECT TRIM("ACCOUNTID") AS acct,
               TRIM("ADDRESSLINE1") AS addr1,
               TRIM("ADDRESSLINE2") AS addr2,
               "ID_CITY" AS id_city,
               TRIM("POSTALCODE") AS zip
        FROM analytics."CUSTADT"
        WHERE import_is_deleted = false
    """)
    for ma in mail_addrs:
        if ma["acct"] not in addr_map and ma["addr1"]:
            city_info_m = city_map.get(ma["id_city"], {})
            addr_map[ma["acct"]] = {
                "addr1": ma["addr1"], "addr2": ma["addr2"],
                "city": city_info_m.get("city"), "state": city_info_m.get("state"),
                "zip": ma["zip"],
            }

    # Sell type → delivery method
    sell_map = {
        "1": "Delivery",
        "2": "Delivery",
        "3": "Will Call",
        "4": "Delivery",
        "5": "Delivery",
        "6": "Will Call",
    }

    # -- Query all active customers from BRATTT --
    customers = query(conn, """
        SELECT TRIM("CMACCT") AS acct,
               TRIM("CMDBA") AS dba,
               TRIM("CMLNAM") AS legal_name,
               TRIM("CMSTS") AS status,
               TRIM("CMTERM"::text) AS term,
               TRIM("CMPHON"::text) AS phone,
               TRIM("CMBSM1") AS salesman,
               TRIM("CMCOMP"::text) AS company,
               TRIM("CMWHSE"::text) AS warehouse,
               TRIM("CMCLAS") AS class_code,
               TRIM("CMPRCD") AS price_code,
               TRIM("CMBRT") AS route,
               TRIM("CMTXCD") AS tax_code,
               TRIM("CMCHN"::text) AS chain,
               TRIM("CMCNTY"::text) AS county,
               TRIM("CMZONE") AS zone,
               TRIM("CMSTTP") AS subtype,
               TRIM("CMSELL") AS sell_type,
               "CMONOF"::text AS on_off,
               TRIM("CMLIC#"::text) AS license,
               TRIM("CMFAX#"::text) AS fax,
               TRIM("CMETHO") AS pay_method,
               "CMIDENTITY" AS identity,
               "ID_CITY" AS id_city,
               TRIM("POSTALCODE") AS postal_code,
               TRIM("CMPGRP") AS price_group
        FROM analytics."BRATTT"
        WHERE import_is_deleted = false
          AND TRIM("CMSTS") = 'A'
          AND TRIM("CMDBA") <> ''
        ORDER BY "CMACCT"
    """)

    # -- Clear existing data --
    clear_data_rows(ws)

    # VLOOKUP ranges
    user_max = 1960   # User tab: ~1958 rows, data starts row 3
    pl_max = 526      # Pricelist tab: ~524 rows, data starts row 3

    print(f"  Writing {len(customers)} customers...")
    for i, c in enumerate(customers):
        row = i + 3
        acct = c["acct"]

        # Resolve lookups
        city_info = city_map.get(c["id_city"], {})
        addr = addr_map.get(acct, {})
        on_off = str(c["on_off"] or "").strip()
        rep_name = rep_map.get((c["salesman"] or "").upper())
        term_desc = terms_map.get(c["term"])
        county_name = county_map.get(c["county"])
        chain_desc = chain_map.get(c["chain"])
        pricelist_name = price_code_to_name.get(c["price_code"])

        # Phone formatting
        phone_raw = c["phone"] or ""
        if phone_raw and phone_raw != "0":
            phone_raw = phone_raw.lstrip("1") if len(phone_raw) > 10 else phone_raw
            if len(phone_raw) == 10:
                phone_val = f"({phone_raw[:3]}) {phone_raw[3:6]}-{phone_raw[6:]}"
            else:
                phone_val = phone_raw
        else:
            phone_val = None

        # Physical address: prefer POSACCT, fall back to CITYT
        phys_street = addr.get("addr1") or None
        if addr.get("addr2"):
            phys_street = f"{phys_street}, {addr['addr2']}" if phys_street else addr["addr2"]
        phys_city = addr.get("city") or city_info.get("city")
        phys_state = addr.get("state") or city_info.get("state")
        phys_zip = addr.get("zip") or c["postal_code"]

        # A: Record Type Name
        ws.cell(row=row, column=1, value="Customer")
        style_cell(ws.cell(row=row, column=1))

        # B: Record Type ID (VLOOKUP)
        ws.cell(row=row, column=2,
                value=f"=VLOOKUP(A{row},'Record Types'!$A$3:$B$6,2,FALSE)")
        style_cell(ws.cell(row=row, column=2), is_formula=True)

        # C: Legacy Name (DBA)
        ws.cell(row=row, column=3, value=s(c["dba"]))
        style_cell(ws.cell(row=row, column=3))

        # D: Customer Name
        ws.cell(row=row, column=4, value=s(c["dba"]))
        style_cell(ws.cell(row=row, column=4))

        # E: Mapping Key (Record Number now at col AR=44)
        ws.cell(row=row, column=5, value=f'=CONCAT("CUST-",AR{row})')
        style_cell(ws.cell(row=row, column=5), is_formula=True)

        # F: Legal Name
        ws.cell(row=row, column=6, value=s(c["legal_name"]))
        style_cell(ws.cell(row=row, column=6))

        # G: Sales Rep Name
        ws.cell(row=row, column=7, value=rep_name)
        style_cell(ws.cell(row=row, column=7))

        # H: Pricelist Name
        ws.cell(row=row, column=8, value=pricelist_name)
        style_cell(ws.cell(row=row, column=8))

        # I-M: Mailing Address (formulas = Physical)
        ws.cell(row=row, column=9, value=f"=N{row}")
        style_cell(ws.cell(row=row, column=9), is_formula=True)
        ws.cell(row=row, column=10, value=f"=O{row}")
        style_cell(ws.cell(row=row, column=10), is_formula=True)
        ws.cell(row=row, column=11, value=f"=P{row}")
        style_cell(ws.cell(row=row, column=11), is_formula=True)
        ws.cell(row=row, column=12, value=f"=Q{row}")
        style_cell(ws.cell(row=row, column=12), is_formula=True)
        ws.cell(row=row, column=13, value=f"=R{row}")
        style_cell(ws.cell(row=row, column=13), is_formula=True)

        # N: Physical Street
        ws.cell(row=row, column=14, value=phys_street)
        style_cell(ws.cell(row=row, column=14))

        # O: Physical City
        ws.cell(row=row, column=15, value=phys_city)
        style_cell(ws.cell(row=row, column=15))

        # P: Physical State
        ws.cell(row=row, column=16, value=phys_state)
        style_cell(ws.cell(row=row, column=16))

        # Q: Physical Zip Code
        ws.cell(row=row, column=17, value=phys_zip)
        style_cell(ws.cell(row=row, column=17))

        # R: Physical Country
        ws.cell(row=row, column=18, value="US")
        style_cell(ws.cell(row=row, column=18))

        # S: Region (county name)
        ws.cell(row=row, column=19, value=county_name)
        style_cell(ws.cell(row=row, column=19))

        # T: Type
        ws.cell(row=row, column=20, value="Customer")
        style_cell(ws.cell(row=row, column=20))

        # U: Stage
        ws.cell(row=row, column=21, value="Customer")
        style_cell(ws.cell(row=row, column=21))

        # V: Payment Terms (mapped to standard Net 0/15/30/60/COD)
        term_code = c["term"] or ""
        ws.cell(row=row, column=22, value=terms_map.get(term_code))
        style_cell(ws.cell(row=row, column=22))

        # W: Payment Method (from CMETHO, with EDI/FinTech override from term code)
        pay_method = term_method_map.get(term_code) or method_map.get(c["pay_method"])
        ws.cell(row=row, column=23, value=pay_method)
        style_cell(ws.cell(row=row, column=23))

        # X: Delivery Method
        ws.cell(row=row, column=24, value=sell_map.get(c["sell_type"]))
        style_cell(ws.cell(row=row, column=24))

        # Y: Consignment Items
        ws.cell(row=row, column=25, value=None)
        style_cell(ws.cell(row=row, column=25))

        # Z: Phone
        ws.cell(row=row, column=26, value=phone_val)
        style_cell(ws.cell(row=row, column=26))

        # AA: Website
        ws.cell(row=row, column=27, value=None)
        style_cell(ws.cell(row=row, column=27))

        # AB: Status
        ws.cell(row=row, column=28, value="Active")
        style_cell(ws.cell(row=row, column=28))

        # AC: Customer Number
        ws.cell(row=row, column=29, value=acct)
        style_cell(ws.cell(row=row, column=29))

        # AD: Premise Type
        ws.cell(row=row, column=30, value=onoff_map.get(on_off))
        style_cell(ws.cell(row=row, column=30))

        # AE: Type (SubType)
        ws.cell(row=row, column=31, value=s(c["subtype"]))
        style_cell(ws.cell(row=row, column=31))

        # AF: ABC Number (license)
        lic = c["license"] or "0"
        ws.cell(row=row, column=32, value=lic if lic != "0" else None)
        style_cell(ws.cell(row=row, column=32))

        # AG: EFT Provider
        ws.cell(row=row, column=33, value=None)
        style_cell(ws.cell(row=row, column=33))

        # AH: EFT Customer ID
        ws.cell(row=row, column=34, value=None)
        style_cell(ws.cell(row=row, column=34))

        # AI: Tax Exempt (CMTXCD: N = not taxable = exempt)
        tax_exempt = (c["tax_code"] or "").upper() == "N"
        ws.cell(row=row, column=35, value=tax_exempt)
        style_cell(ws.cell(row=row, column=35))

        # AJ: Sales Rep Key (VLOOKUP G → User tab S=UserName, T=MappingKey)
        ws.cell(row=row, column=36,
                value=f'=IF(ISNA(VLOOKUP(G{row},User!$S$3:$T${user_max},2,FALSE)),"",VLOOKUP(G{row},User!$S$3:$T${user_max},2,FALSE))')
        style_cell(ws.cell(row=row, column=36), is_formula=True)

        # AK: Pricelist Key (VLOOKUP H → Pricelist A=Name, P=MappingKey)
        ws.cell(row=row, column=37,
                value=f'=IF(ISNA(VLOOKUP(H{row},Pricelist!$A$3:$P${pl_max},16,FALSE)),"",VLOOKUP(H{row},Pricelist!$A$3:$P${pl_max},16,FALSE))')
        style_cell(ws.cell(row=row, column=37), is_formula=True)

        # AL-AQ: New VIP reference columns (38-43)
        # AL: VIP On/Off Code
        ws.cell(row=row, column=38, value=on_off if on_off else None)
        style_cell(ws.cell(row=row, column=38))

        # AM: VIP Route Code
        ws.cell(row=row, column=39, value=s(c["route"]))
        style_cell(ws.cell(row=row, column=39))

        # AN: VIP Chain Code
        ws.cell(row=row, column=40, value=s(c["chain"]))
        style_cell(ws.cell(row=row, column=40))

        # AO: VIP Class Code
        ws.cell(row=row, column=41, value=s(c["class_code"]))
        style_cell(ws.cell(row=row, column=41))

        # AP: VIP Salesman Code
        ws.cell(row=row, column=42, value=s(c["salesman"]))
        style_cell(ws.cell(row=row, column=42))

        # AQ: VIP County Code
        ws.cell(row=row, column=43, value=s(c["county"]))
        style_cell(ws.cell(row=row, column=43))

        # AR: Record Number (col 44)
        ws.cell(row=row, column=44, value=i + 1)
        style_cell(ws.cell(row=row, column=44))

    # Summary stats
    with_addr = sum(1 for c in customers if c["acct"] in addr_map)
    with_rep = sum(1 for c in customers if rep_map.get((c["salesman"] or "").upper()))
    with_pl = sum(1 for c in customers if price_code_to_name.get(c["price_code"]))
    with_terms = sum(1 for c in customers if terms_map.get(c["term"] or ""))
    with_method = sum(1 for c in customers
                      if term_method_map.get(c["term"] or "") or method_map.get(c["pay_method"] or ""))
    print(f"  Address coverage: {with_addr}/{len(customers)} ({100*with_addr//len(customers)}%) have street addresses")
    print(f"  Sales rep resolved: {with_rep}/{len(customers)} ({100*with_rep//len(customers)}%)")
    print(f"  Pricelist resolved: {with_pl}/{len(customers)} ({100*with_pl//len(customers)}%)")
    print(f"  Payment terms: {with_terms}/{len(customers)} ({100*with_terms//len(customers)}%)")
    print(f"  Payment method: {with_method}/{len(customers)} ({100*with_method//len(customers)}%)")
    print(f"  ✓ Customer tab: {len(customers)} rows written")
    return len(customers)


# ---------------------------------------------------------------------------
# Tab: Contact
# ---------------------------------------------------------------------------
def populate_contact(conn, wb):
    """Populate the Contact tab from BRATTT buyer names + APVENT vendor contacts."""
    ws = wb["Contact"]
    print("  Populating Contact tab...")

    # -- Query customer contacts (buyer names from BRATTT) --
    cust_contacts = query(conn, """
        SELECT TRIM("CMACCT") AS acct,
               TRIM("CMDBA") AS dba,
               TRIM("CMBUYR") AS buyer,
               TRIM("CMPHON"::text) AS phone
        FROM analytics."BRATTT"
        WHERE import_is_deleted = false
          AND TRIM("CMSTS") = 'A'
          AND TRIM("CMBUYR") <> ''
        ORDER BY "CMACCT"
    """)

    # -- Query vendor contacts (from APVENT with emails) --
    vendor_contacts = query(conn, """
        SELECT TRIM(v."VRVEND"::text) AS vendor_id,
               TRIM(v."VRNAME") AS vendor_name,
               TRIM(v."VREMAIL") AS email,
               TRIM(v."VRPHON"::text) AS phone
        FROM analytics."APVENT" v
        WHERE v.import_is_deleted = false
          AND TRIM(v."VREMAIL") <> ''
        ORDER BY v."VRVEND"
    """)

    # -- Read Customer tab to build mapping: acct -> row for VLOOKUP --
    # Customer tab col AC = Customer Number (CMACCT), col D = Customer Name
    cust_ws = wb["Customer"]
    cust_name_map = {}
    for row in cust_ws.iter_rows(min_row=3, max_row=cust_ws.max_row, values_only=False):
        acct_val = row[28].value  # col AC (0-indexed = 28) = Customer Number
        name_val = row[3].value   # col D = Customer Name
        if acct_val:
            cust_name_map[str(acct_val).strip()] = str(name_val).strip() if name_val else ""

    # -- Read Supplier tab to build mapping: vendor_id -> supplier name --
    sup_ws = wb["Supplier"]
    sup_name_map = {}
    for row in sup_ws.iter_rows(min_row=3, max_row=sup_ws.max_row, values_only=False):
        vid = row[19].value   # col T (0-indexed = 19) = VIP Vendor Code
        sname = row[1].value  # col B = Name
        if vid:
            sup_name_map[str(vid).strip()] = str(sname).strip() if sname else ""

    # -- Helper: parse buyer name into first/last --
    def parse_name(buyer):
        """Split a buyer name into (first, last). Handles various patterns."""
        if not buyer:
            return ("", "")
        name = buyer.strip()
        # Skip numeric-only values (junk data)
        if name.isdigit():
            return ("", "")
        # Handle HTML entities like &#44; (comma)
        name = name.replace("&#44;", ",")
        # For names with separators (& / OR), take only the first person
        for sep in [" & ", " AND ", " / ", "/", " OR "]:
            if sep in name.upper():
                idx = name.upper().index(sep)
                name = name[:idx].strip()
                break
        # Split into first/last
        parts = name.split()
        if len(parts) == 0:
            return ("", "")
        elif len(parts) == 1:
            return (parts[0].title(), "")
        else:
            return (parts[0].title(), " ".join(parts[1:]).title())

    # -- Helper: format phone --
    def fmt_phone(p):
        if not p:
            return None
        digits = "".join(c for c in str(p) if c.isdigit())
        if len(digits) == 10:
            return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
        if len(digits) == 11 and digits[0] == "1":
            return f"{digits[1:4]}-{digits[4:7]}-{digits[7:]}"
        return p if digits else None

    # -- Clear existing data --
    clear_data_rows(ws)

    # Customer tab max row for VLOOKUP
    cust_max = cust_ws.max_row
    # Supplier tab max row for VLOOKUP
    sup_max = sup_ws.max_row

    rec_num = 0

    # -- Write customer contacts --
    print(f"  Writing {len(cust_contacts)} customer contacts...")
    for i, c in enumerate(cust_contacts):
        row = rec_num + 3
        first, last = parse_name(c["buyer"])
        if not first and not last:
            continue  # skip junk entries

        acct = c["acct"]
        cust_name = cust_name_map.get(acct, c["dba"])
        phone = fmt_phone(c["phone"])

        # A: Customer Name
        cell = ws.cell(row=row, column=1, value=cust_name)
        style_cell(cell)
        # B: Supplier Name (blank for customer contacts)
        style_cell(ws.cell(row=row, column=2, value=None))
        # C: First Name
        cell = ws.cell(row=row, column=3, value=first)
        style_cell(cell)
        # D: Last Name
        cell = ws.cell(row=row, column=4, value=last if last else None)
        style_cell(cell)
        # E: Mobile Phone
        cell = ws.cell(row=row, column=5, value=phone)
        style_cell(cell)
        # F: Email (not available for customers)
        style_cell(ws.cell(row=row, column=6, value=None))
        # G: Title (not available)
        style_cell(ws.cell(row=row, column=7, value=None))
        # H: Department (not available)
        style_cell(ws.cell(row=row, column=8, value=None))
        # I: Description
        style_cell(ws.cell(row=row, column=9, value=None))
        # J: Preferred Contact
        cell = ws.cell(row=row, column=10, value="TRUE")
        style_cell(cell)
        # K: HasOptedOutOfEmail
        style_cell(ws.cell(row=row, column=11, value=None))
        # L: HasOptedOutOfFax
        style_cell(ws.cell(row=row, column=12, value=None))
        # M: DoNotCall
        style_cell(ws.cell(row=row, column=13, value=None))
        # N: Contact Name (formula)
        cell = ws.cell(row=row, column=14,
                       value=f'=C{row}& " " &D{row}')
        style_cell(cell, is_formula=True)
        # O: Mapping Key (formula)
        cell = ws.cell(row=row, column=15,
                       value=f'=CONCAT("CON-",U{row})')
        style_cell(cell, is_formula=True)
        # P: Customer Key — VLOOKUP Customer Name against Customer tab col D->E
        cell = ws.cell(row=row, column=16,
                       value=f'=IF(A{row}="","",IF(ISNA(VLOOKUP(A{row},Customer!$D$3:$E${cust_max},2,FALSE)),"",VLOOKUP(A{row},Customer!$D$3:$E${cust_max},2,FALSE)))')
        style_cell(cell, is_formula=True)
        # Q: Supplier Key (blank for customer contacts)
        style_cell(ws.cell(row=row, column=17, value=None))
        # R: Vendor Key (blank)
        style_cell(ws.cell(row=row, column=18, value=None))
        # S: DC Key (blank)
        style_cell(ws.cell(row=row, column=19, value=None))
        # T: Account Id (formula — picks first non-blank key)
        cell = ws.cell(row=row, column=20,
                       value=f'=IF(P{row}<>"",P{row},IF(Q{row}<>"",Q{row},IF(R{row}<>"",R{row},S{row})))')
        style_cell(cell, is_formula=True)
        # U: Record Number
        rec_num += 1
        cell = ws.cell(row=row, column=21, value=rec_num)
        style_cell(cell)
        # V: True/False
        style_cell(ws.cell(row=row, column=22, value=None))

    cust_count = rec_num

    # -- Write vendor contacts --
    print(f"  Writing {len(vendor_contacts)} vendor contacts...")
    for c in vendor_contacts:
        row = rec_num + 3
        vendor_id = c["vendor_id"]
        sup_name = sup_name_map.get(vendor_id, c["vendor_name"])
        email = c["email"]
        phone = fmt_phone(c["phone"])

        # Parse vendor name — use as contact name if no separate contact
        # APVENT VRNAME is the company name, not a person name
        # We'll put vendor_name as description and leave first/last from vendor_name
        vname = (c["vendor_name"] or "").strip()

        # A: Customer Name (blank for vendor contacts)
        style_cell(ws.cell(row=row, column=1, value=None))
        # B: Supplier Name
        cell = ws.cell(row=row, column=2, value=sup_name)
        style_cell(cell)
        # C: First Name (vendor name as-is, not a person)
        cell = ws.cell(row=row, column=3, value=vname.title() if vname else None)
        style_cell(cell)
        # D: Last Name
        style_cell(ws.cell(row=row, column=4, value=None))
        # E: Mobile Phone
        cell = ws.cell(row=row, column=5, value=phone)
        style_cell(cell)
        # F: Email
        cell = ws.cell(row=row, column=6, value=email)
        style_cell(cell)
        # G-I: blank
        for col in range(7, 10):
            style_cell(ws.cell(row=row, column=col, value=None))
        # J: Preferred Contact
        cell = ws.cell(row=row, column=10, value="TRUE")
        style_cell(cell)
        # K-M: blank
        for col in range(11, 14):
            style_cell(ws.cell(row=row, column=col, value=None))
        # N: Contact Name
        cell = ws.cell(row=row, column=14,
                       value=f'=C{row}& " " &D{row}')
        style_cell(cell, is_formula=True)
        # O: Mapping Key
        cell = ws.cell(row=row, column=15,
                       value=f'=CONCAT("CON-",U{row})')
        style_cell(cell, is_formula=True)
        # P: Customer Key (blank)
        style_cell(ws.cell(row=row, column=16, value=None))
        # Q: Supplier Key — VLOOKUP Supplier Name against Supplier tab col B->D
        cell = ws.cell(row=row, column=17,
                       value=f'=IF(B{row}="","",IF(ISNA(VLOOKUP(B{row},Supplier!$B$3:$D${sup_max},3,FALSE)),"",VLOOKUP(B{row},Supplier!$B$3:$D${sup_max},3,FALSE)))')
        style_cell(cell, is_formula=True)
        # R-S: blank
        style_cell(ws.cell(row=row, column=18, value=None))
        style_cell(ws.cell(row=row, column=19, value=None))
        # T: Account Id
        cell = ws.cell(row=row, column=20,
                       value=f'=IF(P{row}<>"",P{row},IF(Q{row}<>"",Q{row},IF(R{row}<>"",R{row},S{row})))')
        style_cell(cell, is_formula=True)
        # U: Record Number
        rec_num += 1
        cell = ws.cell(row=row, column=21, value=rec_num)
        style_cell(cell)
        # V: blank
        style_cell(ws.cell(row=row, column=22, value=None))

    vendor_count = rec_num - cust_count

    print(f"  Customer contacts: {cust_count}")
    print(f"  Vendor contacts: {vendor_count}")
    print(f"  ✓ Contact tab: {rec_num} total rows written")
    return rec_num


# ---------------------------------------------------------------------------
# Tab: Pickpath
# ---------------------------------------------------------------------------
def populate_pickpath(conn, wb):
    """Populate the Pickpath tab from LOCMAST (warehouse location/pick sequence)."""

    # Create or get the Pickpath sheet
    if "Pickpath" in wb.sheetnames:
        ws = wb["Pickpath"]
    else:
        ws = wb.create_sheet("Pickpath")
    print("  Populating Pickpath tab...")

    # -- Headers (row 1 = display names, row 2 = API/VIP field names) --
    headers_row1 = [
        "Warehouse Code", "Warehouse Name", "Path Sequence", "Location Code",
        "Location Desc", "Location Sequence", "Count Sequence", "Pick Flag",
        "Area", "Status", "Available", "Zone Coord X", "Zone Coord Y",
        "Zone Coord Z", "Pick Type", "Pick Qty", "Replenish", "Staging",
        "Threshold", "Min", "Move Max", "Move Min", "Move MOV",
        "Pallet", "Bond", "Record Number",
    ]
    headers_row2 = [
        "LCWHSE", "warehouse_name", "path_sequence", "LCLOC",
        "LCDESC", "LCLOCSEQ", "LCCNTSEQ", "LCPICK",
        "LCAREA", "LCSTAT", "LCAVAIL", "LCXCOR", "LCYCOR",
        "LCZCOR", "LCPTYP", "LCPQTY", "LCREPL", "LCSTAG",
        "LCTHRS", "LCMIN", "LCMMAX", "LCMMIN", "LCMMOV",
        "LCPALL", "LCBOND", "IDENTITY",
    ]

    # Write headers
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    sub_font = Font(name="Calibri", size=10, italic=True)

    for col_idx, (h1, h2) in enumerate(zip(headers_row1, headers_row2), 1):
        c1 = ws.cell(row=1, column=col_idx, value=h1)
        c1.fill = header_fill
        c1.font = header_font
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=2, column=col_idx, value=h2)
        c2.fill = sub_fill
        c2.font = sub_font
        c2.alignment = Alignment(vertical="center")

    # -- Warehouse name lookup --
    wh_rows = query(conn, """
        SELECT TRIM("WAREHOUSE"::text) AS wh, TRIM("WAREHOUSE_NAME"::text) AS name
        FROM analytics."HDRWHSET"
        WHERE import_is_deleted = false
    """)
    wh_map = {r["wh"]: r["name"] for r in wh_rows}

    # -- Query all LOCMAST rows --
    locations = query(conn, """
        SELECT "IDENTITY" AS id,
               TRIM("LCWHSE"::text) AS wh,
               TRIM("LCLOC") AS loc,
               TRIM("LCDESC") AS descr,
               TRIM("LCLOCSEQ"::text) AS loc_seq,
               TRIM("LCCNTSEQ"::text) AS cnt_seq,
               TRIM("LCPICK") AS pick,
               TRIM("LCAREA") AS area,
               TRIM("LCSTAT") AS stat,
               TRIM("LCAVAIL") AS avail,
               "LCXCOR" AS x, "LCYCOR" AS y, "LCZCOR" AS z,
               TRIM("LCPTYP") AS ptyp,
               "LCPQTY" AS pqty,
               "LCREPL" AS repl,
               TRIM("LCSTAG") AS stag,
               "LCTHRS" AS thrs,
               "LCMIN" AS min_val,
               "LCMMAX" AS mmax,
               "LCMMIN" AS mmin,
               "LCMMOV" AS mmov,
               TRIM("LCPALL") AS pall,
               TRIM("LCBOND") AS bond
        FROM analytics."LOCMAST"
        WHERE import_is_deleted = false
        ORDER BY "LCWHSE", "LCLOCSEQ", "LCCNTSEQ", "LCLOC"
    """)

    # -- Clear existing data rows --
    clear_data_rows(ws)

    # -- Compute per-warehouse path sequence (1, 2, 3... within each warehouse) --
    wh_seq = {}
    for loc in locations:
        wh = loc["wh"]
        wh_seq[wh] = wh_seq.get(wh, 0) + 1
        loc["_path_seq"] = wh_seq[wh]

    # -- Write data --
    print(f"  Writing {len(locations)} location rows...")
    for i, loc in enumerate(locations):
        row = i + 3
        wh = loc["wh"]

        vals = [
            wh,                            # A: Warehouse Code
            wh_map.get(wh, ""),            # B: Warehouse Name
            loc["_path_seq"],               # C: Path Sequence
            loc["loc"],                     # D: Location Code
            loc["descr"],                   # E: Location Desc
            loc["loc_seq"] or None,         # F: Location Sequence
            loc["cnt_seq"] or None,         # G: Count Sequence
            loc["pick"],                    # H: Pick Flag
            loc["area"],                    # I: Area
            loc["stat"],                    # J: Status
            loc["avail"],                   # K: Available
            loc["x"],                       # L: Zone Coord X
            loc["y"],                       # M: Zone Coord Y
            loc["z"],                       # N: Zone Coord Z
            loc["ptyp"],                    # O: Pick Type
            loc["pqty"],                    # P: Pick Qty
            loc["repl"],                    # Q: Replenish
            loc["stag"],                    # R: Staging
            loc["thrs"],                    # S: Threshold
            loc["min_val"],                 # T: Min
            loc["mmax"],                    # U: Move Max
            loc["mmin"],                    # V: Move Min
            loc["mmov"],                    # W: Move MOV
            loc["pall"],                    # X: Pallet
            loc["bond"],                    # Y: Bond
            i + 1,                          # Z: Record Number
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            style_cell(cell)

    # Summary
    pick_y = sum(1 for loc in locations if loc["pick"] == "Y")
    pick_n = sum(1 for loc in locations if loc["pick"] == "N")
    wh_counts = {}
    for loc in locations:
        wh_counts[loc["wh"]] = wh_counts.get(loc["wh"], 0) + 1
    print(f"  Pick locations: {pick_y}, Non-pick: {pick_n}")
    print(f"  Warehouses: {len(wh_counts)}")
    for wh in sorted(wh_counts.keys()):
        print(f"    WH {wh} ({wh_map.get(wh, '?')}): {wh_counts[wh]} locations")
    print(f"  ✓ Pickpath tab: {len(locations)} rows written")
    return len(locations)


# ---------------------------------------------------------------------------
# Tab: Pricelist Item
# ---------------------------------------------------------------------------
def populate_pricelist_item(conn, wb):
    """Populate the Pricelist Item tab from DPMASTT + DPMAST1T deal pricing."""
    ws = wb["Pricelist Item"]
    print("  Populating Pricelist Item tab...")

    # -- Build pricelist lookups from existing Pricelist tab --
    ws_pl = wb["Pricelist"]
    pl_max_row = ws_pl.max_row  # 526 (2 header + 524 data)

    # Map price code → pricelist name, price group → pricelist name
    prcd_to_name = {}
    pgrp_to_name = {}
    for row in range(3, pl_max_row + 1):
        name = ws_pl.cell(row=row, column=1).value   # col A: Name
        pcode = ws_pl.cell(row=row, column=7).value   # col G: VIP Price Code
        pgroup = ws_pl.cell(row=row, column=8).value   # col H: VIP Price Group
        if name:
            name = str(name).strip()
            if pcode:
                prcd_to_name[str(pcode).strip()] = name
            if pgroup:
                pgrp_to_name[str(pgroup).strip()] = name

    print(f"  Pricelist lookups: {len(prcd_to_name)} price codes, {len(pgrp_to_name)} price groups")

    # -- Build item name lookup from ITEMT --
    items = query(conn, """
        SELECT TRIM("ITEM_CODE") AS item_code,
               TRIM("PACKAGE_DESCRIPTION") AS name
        FROM analytics."ITEMT"
        WHERE import_is_deleted = false
          AND TRIM("ITEM_STATUS") IN ('A', 'S')
    """)
    item_name_map = {i["item_code"]: i["name"] for i in items}
    active_items = set(item_name_map.keys())
    print(f"  Active items: {len(active_items)}")

    # -- Query deal pricing: price code items --
    prcd_rows = query(conn, """
        SELECT TRIM(d."DPITEM") AS item_code,
               TRIM(d."DPPRCD") AS plist_key,
               MAX(d1."FRONTLINEPRICE") AS case_price
        FROM analytics."DPMASTT" d
        JOIN analytics."DPMAST1T" d1 ON d."DPIDENTITY" = d1."ID_DPMAST"
        WHERE TRIM(d."DPPRCD") != ''
          AND d."DPEDAT" >= 20250101
          AND d.import_is_deleted = false
        GROUP BY TRIM(d."DPITEM"), TRIM(d."DPPRCD")
    """)

    # -- Query deal pricing: price group items --
    pgrp_rows = query(conn, """
        SELECT TRIM(d."DPITEM") AS item_code,
               TRIM(d."DPPGRP") AS plist_key,
               MAX(d1."FRONTLINEPRICE") AS case_price
        FROM analytics."DPMASTT" d
        JOIN analytics."DPMAST1T" d1 ON d."DPIDENTITY" = d1."ID_DPMAST"
        WHERE TRIM(d."DPPGRP") != ''
          AND d."DPEDAT" >= 20250101
          AND d.import_is_deleted = false
        GROUP BY TRIM(d."DPITEM"), TRIM(d."DPPGRP")
    """)

    # -- Build output rows: filter to active items, non-zero prices, known pricelists --
    output = []
    for r in prcd_rows:
        if r["item_code"] not in active_items:
            continue
        if not r["case_price"] or float(r["case_price"]) <= 0:
            continue
        plist_name = prcd_to_name.get(r["plist_key"])
        if not plist_name:
            continue
        output.append({
            "item_name": item_name_map[r["item_code"]],
            "pricelist": plist_name,
            "case_price": float(r["case_price"]),
        })

    for r in pgrp_rows:
        if r["item_code"] not in active_items:
            continue
        if not r["case_price"] or float(r["case_price"]) <= 0:
            continue
        plist_name = pgrp_to_name.get(r["plist_key"])
        if not plist_name:
            continue
        output.append({
            "item_name": item_name_map[r["item_code"]],
            "pricelist": plist_name,
            "case_price": float(r["case_price"]),
        })

    # Sort by pricelist name then item name
    output.sort(key=lambda x: (x["pricelist"], x["item_name"]))

    print(f"  Deal pricing rows: {len(output)} (from {len(prcd_rows)} prcd + {len(pgrp_rows)} pgrp)")

    # -- Clear existing data --
    clear_data_rows(ws)

    # -- Write data rows --
    # Product tab has data rows 3 to 2746
    # Pricelist tab has data rows 3 to 526
    for i, rec in enumerate(output):
        row = i + 3

        # A: Item Name
        ws.cell(row=row, column=1, value=rec["item_name"])
        style_cell(ws.cell(row=row, column=1))

        # B: Pricelist Name
        ws.cell(row=row, column=2, value=rec["pricelist"])
        style_cell(ws.cell(row=row, column=2))

        # C: Case Price (discounted/deal price)
        ws.cell(row=row, column=3, value=rec["case_price"])
        style_cell(ws.cell(row=row, column=3))

        # D: Default Price (formula: VLOOKUP item key in Product D:J, col 7 = J = Default Price)
        ws.cell(row=row, column=4,
                value=f'=IFERROR(VLOOKUP(G{row},Product!$D:$J,7,FALSE),"")')
        style_cell(ws.cell(row=row, column=4), is_formula=True)

        # E: Discount % (formula)
        ws.cell(row=row, column=5,
                value=f'=IFERROR(IF(D{row}>0,((D{row}-C{row})/D{row})*100,0),0)')
        style_cell(ws.cell(row=row, column=5), is_formula=True)

        # F: Pricelist Mapping Key (VLOOKUP pricelist name → Pricelist col P/16)
        ws.cell(row=row, column=6,
                value=f'=VLOOKUP(B{row},Pricelist!$A$3:$Q${pl_max_row},16,FALSE)')
        style_cell(ws.cell(row=row, column=6), is_formula=True)

        # G: Item Mapping Key (VLOOKUP item name → Product col C:D)
        ws.cell(row=row, column=7,
                value=f'=VLOOKUP(A{row},Product!$C:$D,2,FALSE)')
        style_cell(ws.cell(row=row, column=7), is_formula=True)

    # Stats
    pricelist_counts = {}
    for rec in output:
        pricelist_counts[rec["pricelist"]] = pricelist_counts.get(rec["pricelist"], 0) + 1
    print(f"  Pricelists with items: {len(pricelist_counts)}")
    top_5 = sorted(pricelist_counts.items(), key=lambda x: -x[1])[:5]
    for name, count in top_5:
        print(f"    {name}: {count} items")
    print(f"  ✓ Pricelist Item tab: {len(output)} rows written")
    return len(output)


# ---------------------------------------------------------------------------
# Tab: Route
# ---------------------------------------------------------------------------
def populate_route(conn, wb):
    """Populate the Route tab from HDRROUTET + HDRDRT."""
    ws = wb["Route"]
    print("  Populating Route tab...")

    # -- Query all routes --
    routes = query(conn, """
        SELECT TRIM("ROUTE") AS route_code,
               TRIM("DESCRIPTION") AS description
        FROM analytics."HDRROUTET"
        WHERE import_is_deleted = false
        ORDER BY "ROUTE"
    """)

    # -- Query drivers to match route codes --
    drivers = query(conn, """
        SELECT TRIM("DRIVER") AS driver_code,
               TRIM("DRIVER_NAME") AS driver_name
        FROM analytics."HDRDRT"
        WHERE import_is_deleted = false
    """)
    driver_map = {d["driver_code"]: d["driver_name"] for d in drivers}

    # -- Build driver name in "First Last" format for matching User tab --
    def flip_name(name):
        """Convert 'LAST, FIRST' to 'First Last'."""
        if not name:
            return None
        if "," in name:
            parts = name.split(",", 1)
            last = parts[0].strip().title()
            first = parts[1].strip().title()
            return f"{first} {last}"
        return name.strip().title()

    print(f"  Routes: {len(routes)}, Drivers: {len(driver_map)}")

    # -- Clear existing data --
    clear_data_rows(ws)

    # User tab has data rows 3 to 1960
    user_max = 1960

    for i, r in enumerate(routes):
        row = i + 3
        code = r["route_code"]
        desc = r["description"]

        # Driver name from HDRDRT (route code often = driver code)
        driver_name_raw = driver_map.get(code, desc)
        driver_display = flip_name(driver_name_raw)

        # Determine if route looks like a real driver route or a special route
        is_sample = any(kw in (desc or "").upper() for kw in
                        ["SAMPLE", "OPEN RT", "OVERFLOW", "LOADS FOR", "DRAFT"])
        route_name = f"{code} - {desc}" if desc else code

        # A: Route Name
        ws.cell(row=row, column=1, value=route_name)
        style_cell(ws.cell(row=row, column=1))

        # B: Mapping Key
        ws.cell(row=row, column=2, value=f'=CONCAT("R-",J{row})')
        style_cell(ws.cell(row=row, column=2), is_formula=True)

        # C: Vehicle Name (not available in VIP)
        ws.cell(row=row, column=3, value=None)
        style_cell(ws.cell(row=row, column=3))

        # D: Driver Name (display name for matching)
        ws.cell(row=row, column=4, value=driver_display)
        style_cell(ws.cell(row=row, column=4))

        # E: Active
        ws.cell(row=row, column=5, value=True)
        style_cell(ws.cell(row=row, column=5))

        # F: Day of Week (not in VIP data)
        ws.cell(row=row, column=6, value=None)
        style_cell(ws.cell(row=row, column=6))

        # G: Deliveries Per Day (not in VIP data)
        ws.cell(row=row, column=7, value=None)
        style_cell(ws.cell(row=row, column=7))

        # H: Vehicle lookup (VLOOKUP to VehicleEquipment - empty until that tab is done)
        ws.cell(row=row, column=8,
                value=f'=IFERROR(VLOOKUP(C{row},VehicleEquipment!$C$3:$D$100,2,FALSE),"")')
        style_cell(ws.cell(row=row, column=8), is_formula=True)

        # I: Driver lookup (VLOOKUP driver name to User tab col S/19 → col T/20 Mapping Key)
        ws.cell(row=row, column=9,
                value=f'=IFERROR(VLOOKUP(D{row},User!$S$3:$T${user_max},2,FALSE),"")')
        style_cell(ws.cell(row=row, column=9), is_formula=True)

        # J: Record Number
        if i == 0:
            ws.cell(row=row, column=10, value=1)
        else:
            ws.cell(row=row, column=10, value=f"=J{row - 1}+1")
        style_cell(ws.cell(row=row, column=10))

    sample_routes = sum(1 for r in routes if any(
        kw in (r["description"] or "").upper()
        for kw in ["SAMPLE", "OPEN RT", "OVERFLOW", "LOADS FOR", "DRAFT"]))
    print(f"  Driver routes: {len(routes) - sample_routes}, Special routes: {sample_routes}")
    print(f"  ✓ Route tab: {len(routes)} rows written")
    return len(routes)


# ---------------------------------------------------------------------------
# Tab: AccountRoute
# ---------------------------------------------------------------------------
def populate_account_route(conn, wb):
    """Populate the AccountRoute tab from BRATTT customer-route assignments."""
    ws = wb["AccountRoute"]
    print("  Populating AccountRoute tab...")

    # -- Build route code → route name from Route tab --
    ws_rt = wb["Route"]
    route_max = ws_rt.max_row
    route_code_to_name = {}
    for row in range(3, route_max + 1):
        name = ws_rt.cell(row=row, column=1).value  # col A: Route Name
        if name:
            # Route name format: "CODE - DESCRIPTION"
            code = str(name).split(" - ")[0].strip()
            route_code_to_name[code] = str(name).strip()

    print(f"  Route lookup: {len(route_code_to_name)} routes")

    # -- Query customer-route assignments --
    assignments = query(conn, """
        SELECT TRIM("CMACCT"::text) AS acct,
               TRIM("CMDBA") AS dba,
               TRIM("CMBRT"::text) AS route_code
        FROM analytics."BRATTT"
        WHERE import_is_deleted = false
          AND TRIM("CMSTS") = 'A'
          AND TRIM("CMBRT"::text) <> ''
          AND TRIM("CMDBA") <> ''
        ORDER BY "CMBRT", "CMACCT"
    """)

    # Filter to known routes
    output = []
    unknown_routes = set()
    for a in assignments:
        route_name = route_code_to_name.get(a["route_code"])
        if not route_name:
            unknown_routes.add(a["route_code"])
            continue
        output.append({
            "route_name": route_name,
            "customer_name": a["dba"],
        })

    print(f"  Assignments: {len(output)} (skipped {len(unknown_routes)} unknown route codes)")

    # -- Clear existing data --
    clear_data_rows(ws)

    # Customer tab has data rows 3 to 15339
    cust_max = 15339

    for i, rec in enumerate(output):
        row = i + 3

        # A: Route Name
        ws.cell(row=row, column=1, value=rec["route_name"])
        style_cell(ws.cell(row=row, column=1))

        # B: Customer Name
        ws.cell(row=row, column=2, value=rec["customer_name"])
        style_cell(ws.cell(row=row, column=2))

        # C: Route Key (VLOOKUP route name → Route Mapping Key col B)
        ws.cell(row=row, column=3,
                value=f'=IFERROR(VLOOKUP(A{row},Route!$A$3:$B${route_max},2,FALSE),"")')
        style_cell(ws.cell(row=row, column=3), is_formula=True)

        # D: Customer Key (VLOOKUP customer name → Customer Mapping Key col E)
        ws.cell(row=row, column=4,
                value=f'=IFERROR(VLOOKUP(B{row},Customer!$D$3:$E${cust_max},2,FALSE),"")')
        style_cell(ws.cell(row=row, column=4), is_formula=True)

    # Stats
    route_counts = {}
    for rec in output:
        route_counts[rec["route_name"]] = route_counts.get(rec["route_name"], 0) + 1
    top_5 = sorted(route_counts.items(), key=lambda x: -x[1])[:5]
    for name, count in top_5:
        print(f"    {name}: {count} customers")
    print(f"  ✓ AccountRoute tab: {len(output)} rows written")
    return len(output)


# ---------------------------------------------------------------------------
# Tab: Promotions
# ---------------------------------------------------------------------------
def populate_promotions(conn, wb):
    """Populate the Promotions tab from DISCWKSTT + DISCOUTT + DPMASTT/DPMAST1T."""
    import re
    ws = wb["Promotions"]
    print("  Populating Promotions tab...")

    # -- Query promotion headers (current/recent) --
    promos = query(conn, """
        SELECT TRIM("DISCOUNTID") AS promo_id,
               TRIM("DISCDESC") AS description,
               "STARTDATE" AS start_date,
               "ENDDATE" AS end_date,
               TRIM("DISCOUNTTYPE") AS disc_type
        FROM analytics."DISCWKSTT"
        WHERE import_is_deleted = false
          AND "ENDDATE" >= 20250101
        ORDER BY "DISCOUNTID", "STARTDATE"
    """)
    promo_map = {}
    for p in promos:
        key = (p["promo_id"], int(p["start_date"]), int(p["end_date"]))
        promo_map[key] = p
    print(f"  Promotion headers: {len(promos)}")

    # -- Query promotion-item detail with scope info --
    promo_items = query(conn, """
        SELECT TRIM(d."DISCGRP2") AS promo_id,
               TRIM(d."PRODID") AS item_code,
               d."STARTDATE" AS start_date,
               d."ENDDATE" AS end_date,
               TRIM(d."TYPEID") AS type_id,
               TRIM(d."CODEID") AS code_id
        FROM analytics."DISCOUTT" d
        WHERE d.import_is_deleted = false
          AND d."ENDDATE" >= 20250101
    """)

    # -- Load DPMASTT from CSV (db table is currently empty) --
    # Build lookup: (item, pgrp_or_prcd, whse) → {method, identity, bdat, edat}
    import csv
    dpmastt_csv = os.path.join(os.path.dirname(WORKBOOK_PATH), "csvData", "analytics_DPMASTT.csv")
    dpmastt_map = {}  # key: (item, 'G'|'C', code, whse) → list of {identity, method, bdat, edat}
    with open(dpmastt_csv) as f:
        for row in csv.DictReader(f):
            if row.get("import_is_deleted", "f") == "t":
                continue
            item = row["DPITEM"].strip()
            pgrp = row["DPPGRP"].strip()
            prcd = row["DPPRCD"].strip()
            whse = row["DPWHSE"].strip()
            if whse != "1":
                continue  # filter to warehouse 1 to avoid fan-out
            method = row["DPMETHOD"].strip()
            identity = int(row["DPIDENTITY"])
            bdat = int(row["DPBDAT"])
            edat = int(row["DPEDAT"])
            rec = {"identity": identity, "method": method, "bdat": bdat, "edat": edat,
                   "date_span": edat - bdat}
            if pgrp:
                dpmastt_map.setdefault((item, "G", pgrp), []).append(rec)
            if prcd:
                dpmastt_map.setdefault((item, "C", prcd), []).append(rec)
    print(f"  DPMASTT loaded from CSV: {sum(len(v) for v in dpmastt_map.values())} entries")

    # -- Load DPMAST1T economics from DB (still has 802K rows) --
    dpmast1t_rows = query(conn, """
        SELECT "ID_DPMAST", "SELLINGPRICE01", "REBATEAMOUNT01",
               "REBATEPERCENT", "FRONTLINEPRICE"
        FROM analytics."DPMAST1T"
        WHERE import_is_deleted = false
    """)
    dpmast1t_map = {int(r["ID_DPMAST"]): r for r in dpmast1t_rows}
    print(f"  DPMAST1T loaded from DB: {len(dpmast1t_map)} rows")

    # -- Match promo items to deal economics --
    # For each DISCOUTT row, find DPMASTT entry by (item, type, code) with date overlap
    # Then look up DPMAST1T by DPIDENTITY
    econ_map = {}  # key: (promo_id, item_code, start, end) → economics dict
    for pi in promo_items:
        item_code = pi["item_code"]
        type_id = pi["type_id"]
        code_id = pi.get("code_id", "")
        if not type_id or type_id not in ("G", "C"):
            continue
        dp_key = (item_code, type_id, code_id)
        dp_entries = dpmastt_map.get(dp_key, [])
        if not dp_entries:
            continue
        pi_start = int(pi["start_date"])
        pi_end = int(pi["end_date"])
        # Find best DPMASTT match: date overlap, prefer narrowest span
        best = None
        for dp in dp_entries:
            if dp["bdat"] <= pi_end and dp["edat"] >= pi_start:
                if not best or dp["date_span"] < best["date_span"]:
                    best = dp
        if not best:
            continue
        dp1 = dpmast1t_map.get(best["identity"])
        if not dp1:
            continue
        econ_key = (pi["promo_id"], item_code, pi_start, pi_end)
        econ_map[econ_key] = {
            "method": best["method"],
            "selling_price": dp1["SELLINGPRICE01"],
            "rebate_amount": dp1["REBATEAMOUNT01"],
            "rebate_pct": dp1["REBATEPERCENT"],
            "frontline_price": dp1["FRONTLINEPRICE"],
        }
    print(f"  Deal economics matched: {len(econ_map)} promo-item combos")

    # -- Build item lookup: code → (name, brand_code) --
    items = query(conn, """
        SELECT TRIM("ITEM_CODE") AS item_code,
               TRIM("PACKAGE_DESCRIPTION") AS name,
               TRIM("BRAND_CODE") AS brand_code
        FROM analytics."ITEMT"
        WHERE import_is_deleted = false
          AND TRIM("ITEM_STATUS") IN ('A', 'S')
    """)
    item_name_map = {i["item_code"]: i["name"] for i in items}
    item_brand_map = {i["item_code"]: i["brand_code"] for i in items}
    active_items = set(item_name_map.keys())

    # -- Build brand lookup --
    brands = query(conn, """
        SELECT TRIM("BRAND") AS code, TRIM("BRAND_NAME") AS name
        FROM analytics."BRANDT" WHERE import_is_deleted = false
    """)
    brand_name_map = {b["code"]: b["name"] for b in brands}

    # -- Build price group name lookup (for Type G scope) --
    pgroups = query(conn, """
        SELECT TRIM("PRICEGROUP") AS code, TRIM("DESCRIPTION") AS name
        FROM analytics."HDRPRGRPT" WHERE import_is_deleted = false
    """)
    pgroup_name_map = {g["code"]: g["name"] for g in pgroups}

    # -- Build price code name lookup (for Type C scope) --
    pcodes = query(conn, """
        SELECT TRIM("PRICECODE") AS code, TRIM("DESCRIPTION") AS name
        FROM analytics."HDRPRCODT" WHERE import_is_deleted = false
    """)
    pcode_name_map = {c["code"]: c["name"] for c in pcodes}

    print(f"  Active items: {len(active_items)}, Brands: {len(brand_name_map)}")

    # -- Parse promotion description for quantity --
    def parse_promo_desc(desc):
        if not desc:
            return None, None, "Unknown"
        qty_match = re.search(r'(\d+)(?:/(\d+))?\s*(?:CS|CASE)\b', desc, re.IGNORECASE)
        qty_min = int(qty_match.group(1)) if qty_match else None
        qty_max = int(qty_match.group(2)) if qty_match and qty_match.group(2) else None

        # Determine deal type from description
        desc_upper = desc.upper()
        dollar_match = re.search(r'\$(\d+(?:\.\d+)?)', desc)
        has_dollar = dollar_match is not None
        if "QD" in desc_upper and has_dollar:
            deal_type = "Quantity Discount + $ Off"
        elif "QD" in desc_upper:
            deal_type = "Quantity Discount"
        elif has_dollar and "OFF" in desc_upper:
            deal_type = "$ Off"
        elif has_dollar:
            deal_type = "Fixed Price"
        elif "BUY" in desc_upper and "GET" in desc_upper:
            deal_type = "Buy X Get Y"
        elif "PICK" in desc_upper:
            deal_type = "Mix & Match"
        elif "CONST" in desc_upper:
            deal_type = "Constellation Program"
        elif any(kw in desc_upper for kw in ["DIAMOND", "PLATINUM", "GOLD", "BRONZE"]):
            deal_type = "Tiered Program"
        else:
            deal_type = "Promotion"

        return qty_min, qty_max, deal_type

    # -- Scope description --
    def get_scope(type_id, code_id):
        if type_id == "G":
            grp_name = pgroup_name_map.get(code_id, code_id)
            return "Price Group", grp_name
        elif type_id == "C":
            code_name = pcode_name_map.get(code_id, code_id)
            return "Price Code", code_name
        elif type_id == "A":
            return "All Customers", ""
        return "Unknown", ""

    # -- Format economics based on method --
    # Only show relevant fields per method to avoid confusing zeros
    def format_method(econ):
        if not econ:
            return "", None, None, None, None
        method = econ.get("method", "")
        sell = econ.get("selling_price")
        rebate = econ.get("rebate_amount")
        rebate_pct = econ.get("rebate_pct")
        frontline = econ.get("frontline_price")
        if method == "2":
            # Percentage deal: show frontline + rebate %, suppress rebate $
            return "Billback %", sell, None, rebate_pct, frontline
        else:
            # Dollar-off deal: show selling price + rebate $, suppress % and frontline
            return "Billback $", sell, rebate, None, None

    # -- Match promo items to headers and build output --
    output = []
    unmatched_promos = set()
    econ_matched = 0
    econ_missed = 0
    for pi in promo_items:
        if pi["item_code"] not in active_items:
            continue
        key = (pi["promo_id"], int(pi["start_date"]), int(pi["end_date"]))
        promo = promo_map.get(key)
        if not promo:
            unmatched_promos.add(pi["promo_id"])
            continue

        item_name = item_name_map[pi["item_code"]]
        brand_code = item_brand_map.get(pi["item_code"], "")
        brand_name = brand_name_map.get(brand_code, "")
        qty_min, qty_max, deal_type = parse_promo_desc(promo["description"])
        scope_type, scope_detail = get_scope(pi["type_id"], pi.get("code_id", ""))

        # Look up deal economics
        econ_key = (pi["promo_id"], pi["item_code"],
                    int(pi["start_date"]), int(pi["end_date"]))
        econ = econ_map.get(econ_key)
        method_label, sell_price, rebate_amt, rebate_pct, frontline = format_method(econ)
        if econ:
            econ_matched += 1
        else:
            econ_missed += 1

        start_raw = str(int(promo["start_date"]))
        end_raw = str(int(promo["end_date"]))
        start_fmt = f"{start_raw[:4]}-{start_raw[4:6]}-{start_raw[6:]}" if len(start_raw) == 8 else start_raw
        end_fmt = f"{end_raw[:4]}-{end_raw[4:6]}-{end_raw[6:]}" if len(end_raw) == 8 else end_raw

        output.append({
            "promo_id": promo["promo_id"],
            "promo_desc": promo["description"],
            "item_name": item_name,
            "brand": brand_name,
            "deal_type": deal_type,
            "method": method_label,
            "selling_price": sell_price,
            "rebate_amt": rebate_amt,
            "rebate_pct": rebate_pct,
            "frontline": frontline,
            "qty_min": qty_min,
            "qty_max": qty_max,
            "start_date": start_fmt,
            "end_date": end_fmt,
            "scope_type": scope_type,
            "scope_detail": scope_detail,
        })

    output.sort(key=lambda x: (x["promo_id"] or "", x["brand"] or "", x["item_name"] or ""))

    if unmatched_promos:
        print(f"  Skipped {len(unmatched_promos)} promo IDs with no matching header")
    print(f"  Promotion-item rows: {len(output)}")
    print(f"  Deal economics: {econ_matched} matched, {econ_missed} unmatched")

    # -- Update headers (row 1) --
    headers = [
        "Promotion Name", "Deal Description", "Deal Type", "Item", "Brand",
        "Method", "Selling Price", "Rebate $/Case", "Rebate %", "Frontline Price",
        "Qty Min", "Qty Max", "Start Date", "End Date",
        "Applies To", "Scope Detail", "Promo ID"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(vertical="center")
    # Clear row 2 (template had no API names)
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col, value=None)

    # -- Clear existing data --
    clear_data_rows(ws)

    # -- Write data rows --
    for i, rec in enumerate(output):
        row = i + 3

        # A: Promotion Name (ID + description)
        ws.cell(row=row, column=1, value=f"{rec['promo_id']} - {rec['promo_desc']}")
        style_cell(ws.cell(row=row, column=1))

        # B: Deal Description (the raw promo description)
        ws.cell(row=row, column=2, value=rec["promo_desc"])
        style_cell(ws.cell(row=row, column=2))

        # C: Deal Type (parsed from description)
        ws.cell(row=row, column=3, value=rec["deal_type"])
        style_cell(ws.cell(row=row, column=3))

        # D: Item
        ws.cell(row=row, column=4, value=rec["item_name"])
        style_cell(ws.cell(row=row, column=4))

        # E: Brand
        ws.cell(row=row, column=5, value=rec["brand"])
        style_cell(ws.cell(row=row, column=5))

        # F: Method (Billback $ / Billback %)
        ws.cell(row=row, column=6, value=rec["method"])
        style_cell(ws.cell(row=row, column=6))

        # G: Selling Price (from DPMAST1T.SELLINGPRICE01)
        ws.cell(row=row, column=7, value=rec["selling_price"])
        style_cell(ws.cell(row=row, column=7))

        # H: Rebate $/Case (from DPMAST1T.REBATEAMOUNT01)
        ws.cell(row=row, column=8, value=rec["rebate_amt"])
        style_cell(ws.cell(row=row, column=8))

        # I: Rebate % (from DPMAST1T.REBATEPERCENT)
        ws.cell(row=row, column=9, value=rec["rebate_pct"])
        style_cell(ws.cell(row=row, column=9))

        # J: Frontline Price (from DPMAST1T.FRONTLINEPRICE, used for % deals)
        ws.cell(row=row, column=10, value=rec["frontline"])
        style_cell(ws.cell(row=row, column=10))

        # K: Qty Min (parsed from description)
        ws.cell(row=row, column=11, value=rec["qty_min"])
        style_cell(ws.cell(row=row, column=11))

        # L: Qty Max (parsed from description)
        ws.cell(row=row, column=12, value=rec["qty_max"])
        style_cell(ws.cell(row=row, column=12))

        # M: Start Date
        ws.cell(row=row, column=13, value=rec["start_date"])
        style_cell(ws.cell(row=row, column=13))

        # N: End Date
        ws.cell(row=row, column=14, value=rec["end_date"])
        style_cell(ws.cell(row=row, column=14))

        # O: Applies To (scope type)
        ws.cell(row=row, column=15, value=rec["scope_type"])
        style_cell(ws.cell(row=row, column=15))

        # P: Scope Detail (price group/code name)
        ws.cell(row=row, column=16, value=rec["scope_detail"])
        style_cell(ws.cell(row=row, column=16))

        # Q: Promo ID (for reference/dedup)
        ws.cell(row=row, column=17, value=rec["promo_id"])
        style_cell(ws.cell(row=row, column=17))

    # Stats
    promo_counts = {}
    for rec in output:
        promo_counts[rec["promo_id"]] = promo_counts.get(rec["promo_id"], 0) + 1
    econ_pct = (econ_matched / (econ_matched + econ_missed) * 100) if (econ_matched + econ_missed) > 0 else 0
    deal_types = {}
    for rec in output:
        deal_types[rec["deal_type"]] = deal_types.get(rec["deal_type"], 0) + 1
    methods = {}
    for rec in output:
        m = rec["method"] or "No Match"
        methods[m] = methods.get(m, 0) + 1
    print(f"  Distinct promotions: {len(promo_counts)}")
    print(f"  Economics enrichment: {econ_pct:.1f}%")
    print(f"  Methods:")
    for m, cnt in sorted(methods.items(), key=lambda x: -x[1]):
        print(f"    {m}: {cnt}")
    print(f"  Deal types:")
    for dt, cnt in sorted(deal_types.items(), key=lambda x: -x[1]):
        print(f"    {dt}: {cnt}")
    print(f"  ✓ Promotions tab: {len(output)} rows written")
    return len(output)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
TAB_HANDLERS = {
    "User": populate_user,
    "ParentLocation": populate_parent_location,
    "ChildLocation": populate_child_location,
    "Product": populate_product,
    "Customer": populate_customer,
    "Contact": populate_contact,
    "Pickpath": populate_pickpath,
    "Pricelist Item": populate_pricelist_item,
    "Route": populate_route,
    "AccountRoute": populate_account_route,
    "Promotions": populate_promotions,
}


def main():
    target_tab = sys.argv[1] if len(sys.argv) > 1 else None

    if target_tab and target_tab not in TAB_HANDLERS:
        available = ", ".join(TAB_HANDLERS.keys())
        print(f"Error: Unknown tab '{target_tab}'. Available: {available}")
        sys.exit(1)

    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    print("Connecting to database...")
    conn = get_connection()

    tabs_to_run = {target_tab: TAB_HANDLERS[target_tab]} if target_tab else TAB_HANDLERS

    for tab_name, handler in tabs_to_run.items():
        print(f"\n--- {tab_name} ---")
        handler(conn, wb)

    conn.close()

    print(f"\nSaving workbook...")
    wb.save(WORKBOOK_PATH)
    print(f"Done! Saved to {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
