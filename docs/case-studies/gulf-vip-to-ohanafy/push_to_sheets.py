"""
Push populated tabs from the local Excel workbook to Google Sheets.

Usage:
    python3 push_to_sheets.py [tab_name]

    tab_name: Optional. Push only that tab. If omitted, pushes all populated tabs.
              e.g. "Customer", "Pickpath", "Contact"

Requires:
    - service-account.json in the same directory
    - Google Sheet shared with the service account email
"""

import os
import sys
import time
import openpyxl
import gspread
from decimal import Decimal
from datetime import datetime, date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKBOOK_PATH = os.path.join(SCRIPT_DIR, "Distro - Data Workbook TEMPLATE 3_3.xlsx")
SERVICE_ACCOUNT_PATH = os.path.join(SCRIPT_DIR, "service-account.json")
SHEET_ID = "1j8Uuv_7IGV3r-VZmpCnzVea3GDjPrAtM2xgi3jFGNZQ"

# Tabs we've populated and want to push
PUSHABLE_TABS = [
    "User",
    "Pricelist",
    "Supplier",
    "Suppliers(Item Line)",
    "Brands",
    "ParentLocation",
    "ChildLocation",
    "Product",
    "Customer",
    "Contact",
    "Pickpath",
    "Pricelist Item",
    "Route",
    "AccountRoute",
    "Promotions",
]

# Google Sheets API rate limit: 60 requests/min per user
# batch_update counts as 1 request regardless of size
BATCH_SIZE = 5000  # rows per batch update


def clean_value(val):
    """Convert cell value to something JSON-serializable for Google Sheets."""
    if val is None:
        return ""
    if isinstance(val, Decimal):
        # Convert Decimal to int if it's whole, otherwise float
        if val == int(val):
            return int(val)
        return float(val)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    return str(val)


def push_tab(gc, sh, wb, tab_name):
    """Push a single tab from the Excel workbook to Google Sheets."""
    # Get the Excel worksheet
    if tab_name not in wb.sheetnames:
        print(f"  ✗ Tab '{tab_name}' not found in Excel workbook")
        return False

    xl_ws = wb[tab_name]
    max_row = xl_ws.max_row
    max_col = xl_ws.max_column

    if max_row < 3:
        print(f"  ✗ Tab '{tab_name}' has no data rows")
        return False

    print(f"  {tab_name}: {max_row - 2} data rows, {max_col} columns")

    # Get or create the Google Sheet tab
    try:
        gs_ws = sh.worksheet(tab_name)
    except gspread.WorksheetNotFound:
        print(f"    Creating new sheet '{tab_name}'...")
        gs_ws = sh.add_worksheet(title=tab_name, rows=max_row, cols=max_col)

    # Read all data from Excel
    all_rows = []
    for row in xl_ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col,
                                values_only=True):
        all_rows.append([clean_value(v) for v in row])

    # Resize Google Sheet if needed
    if gs_ws.row_count < max_row or gs_ws.col_count < max_col:
        gs_ws.resize(rows=max(max_row, gs_ws.row_count),
                     cols=max(max_col, gs_ws.col_count))

    # Clear existing data
    gs_ws.clear()

    # Push in batches
    total_rows = len(all_rows)
    for start in range(0, total_rows, BATCH_SIZE):
        end = min(start + BATCH_SIZE, total_rows)
        batch = all_rows[start:end]
        start_row = start + 1  # 1-indexed

        # Use update with range
        cell_range = f"A{start_row}"
        gs_ws.update(batch, cell_range, value_input_option='USER_ENTERED')

        pushed = end
        print(f"    Pushed rows {start + 1}-{end} of {total_rows}")

        # Rate limit pause between batches
        if end < total_rows:
            time.sleep(2)

    print(f"  ✓ {tab_name}: {total_rows} rows pushed")
    return True


def main():
    target_tab = sys.argv[1] if len(sys.argv) > 1 else None

    if target_tab and target_tab not in PUSHABLE_TABS:
        print(f"Available tabs: {', '.join(PUSHABLE_TABS)}")
        sys.exit(1)

    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)

    print("Connecting to Google Sheets...")
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_PATH)
    sh = gc.open_by_key(SHEET_ID)
    print(f"Connected to: {sh.title}\n")

    tabs = [target_tab] if target_tab else PUSHABLE_TABS
    success = 0
    for tab in tabs:
        try:
            if push_tab(gc, sh, wb, tab):
                success += 1
            # Pause between tabs to avoid rate limits
            time.sleep(3)
        except gspread.exceptions.APIError as e:
            if "429" in str(e) or "Quota" in str(e):
                print(f"    Rate limited. Waiting 60s...")
                time.sleep(60)
                if push_tab(gc, sh, wb, tab):
                    success += 1
            else:
                print(f"  ✗ {tab}: API error: {e}")

    wb.close()
    print(f"\nDone! {success}/{len(tabs)} tabs pushed to Google Sheets.")


if __name__ == "__main__":
    main()
