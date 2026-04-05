"""
Skill: data-harmonizer

Ingests customer Excel/CSV files, maps columns to Ohanafy's Salesforce schema
using Claude Sonnet, validates data, and stages cleaned records for human review.

Orchestration flow: intake → sample → map → validate → stage → approve
"""

from __future__ import annotations

import csv
import io
import json
import random
import uuid
from datetime import datetime, timezone
from pathlib import Path

import anthropic
import boto3
import yaml
from pydantic import ValidationError

try:
    from .schema import (
        SF_OBJECTS,
        REQUIRED_FIELDS,
        ColumnMapping,
        Confidence,
        InputSchema,
        MappingResult,
        OutputSchema,
        StagedRecord,
        ValidationIssue,
    )
except ImportError:
    from schema import (
        SF_OBJECTS,
        REQUIRED_FIELDS,
        ColumnMapping,
        Confidence,
        InputSchema,
        MappingResult,
        OutputSchema,
        StagedRecord,
        ValidationIssue,
    )

# --- Constants ---

PROMPT_PATH = Path(__file__).parent.parent.parent / "claude" / "prompts" / "data-mapping-v1.0.yaml"
MAX_SAMPLE_ROWS = 20  # 10 first + 10 random
MAX_PRIOR_MAPPINGS = 5
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}


# --- File Reading ---


def read_file(file_path: str, sheet_name: str | None, header_row: int) -> tuple[list[str], list[dict]]:
    """Read an Excel or CSV file and return (headers, rows as dicts).

    Args:
        file_path: Path to the file.
        sheet_name: Sheet name for xlsx files (None = first sheet).
        header_row: 1-indexed row number for headers.

    Returns:
        Tuple of (column_headers, list_of_row_dicts).
    """
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {SUPPORTED_EXTENSIONS}")

    if ext == ".csv":
        return _read_csv(path, header_row)
    elif ext == ".xlsx":
        return _read_xlsx(path, sheet_name, header_row)

    raise ValueError(f"Unsupported extension: {ext}")


def _read_csv(path: Path, header_row: int) -> tuple[list[str], list[dict]]:
    """Read a CSV file."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        lines = f.readlines()

    # Skip to header row (1-indexed)
    header_idx = header_row - 1
    if header_idx >= len(lines):
        raise ValueError(f"Header row {header_row} exceeds file length ({len(lines)} lines)")

    reader = csv.DictReader(io.StringIO("".join(lines[header_idx:])))
    headers = reader.fieldnames or []
    rows = [row for row in reader]
    return headers, rows


def _read_xlsx(path: Path, sheet_name: str | None, header_row: int) -> tuple[list[str], list[dict]]:
    """Read an Excel file. Requires openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for .xlsx files: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    all_rows = list(rows_iter)
    wb.close()

    if not all_rows or header_row - 1 >= len(all_rows):
        raise ValueError(f"Header row {header_row} exceeds sheet length ({len(all_rows)} rows)")

    headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(all_rows[header_row - 1])]
    data_rows = []
    for row in all_rows[header_row:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        data_rows.append(row_dict)

    return headers, data_rows


# --- Sampling ---


def sample_rows(rows: list[dict], max_samples: int = MAX_SAMPLE_ROWS) -> list[dict]:
    """Sample rows for Claude analysis: first 10 + 10 random from body.

    For files with <=max_samples rows, return all rows.
    """
    if len(rows) <= max_samples:
        return rows

    half = max_samples // 2
    first_rows = rows[:half]
    remaining = rows[half:]
    random_rows = random.sample(remaining, min(half, len(remaining)))
    return first_rows + random_rows


# --- Prior Mappings ---


def get_prior_mappings(customer_id: str, table_name: str) -> list[dict]:
    """Retrieve prior approved mappings from DynamoDB for few-shot context.

    Returns the N most recent approved mappings for this customer.
    """
    try:
        dynamodb = boto3.resource("dynamodb")
        table = dynamodb.Table(table_name)
        response = table.query(
            KeyConditionExpression=boto3.dynamodb.conditions.Key("customer_id").eq(customer_id),
            ScanIndexForward=False,  # newest first
            Limit=MAX_PRIOR_MAPPINGS,
        )
        return response.get("Items", [])
    except Exception:
        # If DynamoDB is not available (e.g., local dev), return empty
        return []


def format_prior_mappings(mappings: list[dict]) -> str:
    """Format prior mappings as text for the prompt."""
    if not mappings:
        return "No prior mappings found for this customer. This is a first-time import."

    lines = ["Previously approved mappings for this customer:"]
    for m in mappings:
        source = m.get("source_column", "?")
        target_obj = m.get("target_object", "?")
        target_field = m.get("target_field", "?")
        lines.append(f"  - \"{source}\" → {target_obj}.{target_field}")
    return "\n".join(lines)


# --- SF Schema Formatting ---


def format_sf_schema() -> str:
    """Format the SF schema for inclusion in the mapping prompt."""
    lines = []
    for obj_name, model_cls in SF_OBJECTS.items():
        required = REQUIRED_FIELDS.get(obj_name, [])
        fields = []
        for field_name, field_info in model_cls.model_fields.items():
            req_marker = " (REQUIRED)" if field_name in required else ""
            desc = field_info.description or ""
            fields.append(f"    - {field_name}: {desc}{req_marker}")

        lines.append(f"  {obj_name}:")
        lines.extend(fields)
        lines.append("")
    return "\n".join(lines)


# --- Claude Mapping ---


def call_claude_mapping(
    headers: list[str],
    sample: list[dict],
    customer_id: str,
    prior_mappings: list[dict],
) -> MappingResult:
    """Call Claude Sonnet to map Excel columns to SF schema."""
    # Load prompt template
    with open(PROMPT_PATH) as f:
        prompt_config = yaml.safe_load(f)

    template = prompt_config["template"]

    # Render prompt
    rendered = template.format(
        sf_schema=format_sf_schema(),
        column_headers=json.dumps(headers),
        sample_rows=json.dumps(sample[:MAX_SAMPLE_ROWS], indent=2, default=str),
        customer_id=customer_id,
        prior_mappings=format_prior_mappings(prior_mappings),
    )

    # Call Claude Sonnet
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{"role": "user", "content": rendered}],
    )

    # Parse response
    response_text = response.content[0].text

    # Extract JSON from response (may be wrapped in ```json blocks)
    json_str = response_text
    if "```json" in json_str:
        json_str = json_str.split("```json")[1].split("```")[0]
    elif "```" in json_str:
        json_str = json_str.split("```")[1].split("```")[0]

    parsed = json.loads(json_str.strip())

    # Build MappingResult
    mappings = []
    for m in parsed.get("mappings", []):
        mappings.append(
            ColumnMapping(
                source_column=m["source_column"],
                target_object=m["target_object"],
                target_field=m["target_field"],
                confidence=Confidence(m["confidence"]),
                rationale=m["rationale"],
                transform=m.get("transform"),
            )
        )

    return MappingResult(
        mappings=mappings,
        unmapped_columns=parsed.get("unmapped_columns", []),
    )


# --- Validation ---


def validate_records(
    rows: list[dict],
    mapping: MappingResult,
) -> tuple[list[StagedRecord], list[ValidationIssue]]:
    """Apply mappings to rows and validate against SF schema.

    Returns (staged_records, validation_issues).
    """
    staged = []
    issues = []

    # Build lookup: source_column -> mapping
    col_map: dict[str, ColumnMapping] = {}
    for m in mapping.mappings:
        if m.confidence != Confidence.LOW:  # skip low-confidence mappings
            col_map[m.source_column] = m

    # Group mappings by target object
    target_objects = set(m.target_object for m in col_map.values())

    for row_idx, row in enumerate(rows, start=2):  # 1-indexed, row 1 is header
        for target_obj in target_objects:
            fields = {}
            original = {}
            obj_mappings = [m for m in col_map.values() if m.target_object == target_obj]

            for m in obj_mappings:
                value = row.get(m.source_column)
                original[m.source_column] = value

                if value is not None and str(value).strip():
                    fields[m.target_field] = str(value).strip()

            # Skip empty rows (no mapped values)
            if not fields:
                continue

            # Validate required fields
            required = REQUIRED_FIELDS.get(target_obj, [])
            for req_field in required:
                if req_field not in fields or not fields[req_field]:
                    issues.append(
                        ValidationIssue(
                            row=row_idx,
                            column=req_field,
                            issue=f"Required field '{req_field}' is missing for {target_obj}",
                            severity="error",
                        )
                    )

            staged.append(
                StagedRecord(
                    source_row=row_idx,
                    target_object=target_obj,
                    fields=fields,
                    original_values=original,
                )
            )

    return staged, issues


# --- Duplicate Detection ---


def check_duplicates(
    staged: list[StagedRecord],
) -> list[ValidationIssue]:
    """Basic in-file duplicate detection based on Name fields."""
    issues = []
    seen: dict[str, list[int]] = {}

    for record in staged:
        name = record.fields.get("Name", "")
        if not name:
            continue
        key = f"{record.target_object}:{name.lower()}"
        if key in seen:
            issues.append(
                ValidationIssue(
                    row=record.source_row,
                    column="Name",
                    issue=f"Possible duplicate: '{name}' also appears in row(s) {seen[key]}",
                    severity="warning",
                )
            )
            seen[key].append(record.source_row)
        else:
            seen[key] = [record.source_row]

    return issues


# --- Logging ---


def log_event(
    table_name: str,
    batch_id: str,
    event_type: str,
    customer_id: str,
    data: dict,
) -> None:
    """Log an event to the harmonizer log table in DynamoDB."""
    try:
        dynamodb = boto3.resource("dynamodb")
        table = dynamodb.Table(table_name)
        now = datetime.now(timezone.utc).isoformat()
        table.put_item(
            Item={
                "batch_id": batch_id,
                "event_key": f"{event_type}#{now}",
                "customer_id": customer_id,
                "timestamp": now,
                "event_type": event_type,
                "data": json.loads(json.dumps(data, default=str)),
            }
        )
    except Exception:
        # Logging failure should not break the skill
        pass


# --- Main Orchestration ---


def run(params: InputSchema) -> OutputSchema:
    """Execute the data-harmonizer skill.

    Flow: intake → sample → map → validate → stage

    Args:
        params: Validated input parameters.

    Returns:
        OutputSchema with mapping results, staged records, and validation issues.
    """
    batch_id = f"batch-{uuid.uuid4().hex[:12]}"
    mappings_table = f"ohanafy-{_get_env()}-mappings"
    log_table = f"ohanafy-{_get_env()}-harmonizer-log"

    try:
        # 1. INTAKE — read the file
        headers, rows = read_file(
            params.file_path,
            params.sheet_name,
            params.header_row,
        )

        if not headers:
            return OutputSchema(
                status="error",
                error={"code": "EMPTY_FILE", "message": "File contains no column headers"},
            )

        if not rows:
            return OutputSchema(
                status="error",
                error={"code": "NO_DATA", "message": "File contains headers but no data rows"},
            )

        # Log upload event
        log_event(log_table, batch_id, "upload", params.customer_id, {
            "file_path": params.file_path,
            "headers": headers,
            "row_count": len(rows),
        })

        # 2. SAMPLE — select rows for Claude analysis
        sample = sample_rows(rows)

        # 3. MAP — call Claude to map columns
        prior = get_prior_mappings(params.customer_id, mappings_table)
        mapping = call_claude_mapping(headers, sample, params.customer_id, prior)
        mapping.source_row_count = len(rows)
        mapping.sample_rows_analyzed = len(sample)

        # Log mapping result
        log_event(log_table, batch_id, "mapping", params.customer_id, {
            "mappings_count": len(mapping.mappings),
            "unmapped_count": len(mapping.unmapped_columns),
            "confidence_breakdown": {
                "high": sum(1 for m in mapping.mappings if m.confidence == Confidence.HIGH),
                "medium": sum(1 for m in mapping.mappings if m.confidence == Confidence.MEDIUM),
                "low": sum(1 for m in mapping.mappings if m.confidence == Confidence.LOW),
            },
        })

        # 4. VALIDATE — apply mappings and check data quality
        staged, issues = validate_records(rows, mapping)
        dupe_issues = check_duplicates(staged)
        issues.extend(dupe_issues)

        # Count errors vs warnings
        error_count = sum(1 for i in issues if i.severity == "error")
        warning_count = sum(1 for i in issues if i.severity == "warning")

        # Determine status
        if error_count > 0 or mapping.unmapped_columns:
            status = "needs_review"
        else:
            status = "success"

        stats = {
            "total_rows": len(rows),
            "mapped_rows": len(staged),
            "error_rows": error_count,
            "warning_rows": warning_count,
            "unmapped_columns": len(mapping.unmapped_columns),
            "high_confidence_mappings": sum(1 for m in mapping.mappings if m.confidence == Confidence.HIGH),
            "medium_confidence_mappings": sum(1 for m in mapping.mappings if m.confidence == Confidence.MEDIUM),
        }

        # Log validation result
        log_event(log_table, batch_id, "validation", params.customer_id, stats)

        return OutputSchema(
            status=status,
            batch_id=batch_id,
            mapping=mapping,
            staged_records=staged if not params.dry_run else [],
            validation_issues=issues,
            stats=stats,
        )

    except ValueError as e:
        return OutputSchema(
            status="error",
            error={"code": "VALIDATION_ERROR", "message": str(e)},
        )
    except json.JSONDecodeError as e:
        return OutputSchema(
            status="error",
            error={"code": "MAPPING_PARSE_ERROR", "message": f"Failed to parse Claude mapping response: {e}"},
        )
    except anthropic.APIError as e:
        return OutputSchema(
            status="error",
            error={"code": "CLAUDE_API_ERROR", "message": str(e)},
        )
    except Exception as e:
        return OutputSchema(
            status="error",
            error={"code": "UNEXPECTED_ERROR", "message": str(e)},
        )


def _get_env() -> str:
    """Get current environment from env var or default to dev."""
    import os
    return os.environ.get("OHANAFY_ENV", "dev")
