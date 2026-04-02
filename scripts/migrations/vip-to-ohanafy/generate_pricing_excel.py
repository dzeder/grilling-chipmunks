#!/usr/bin/env python3
"""
Generate VIP Pricing Data Export Excel workbook.

Connects to the Gulf Distributing VIP database and produces a multi-sheet
Excel workbook for pricing analysis and migration planning.

Output: VIP_Pricing_Data_Export.xlsx
"""

import os
import sys
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Database connection
# ---------------------------------------------------------------------------
DB_CONFIG = {
    "host": os.getenv("PGHOST", "gulfstream-db2-data.postgres.database.azure.com"),
    "port": int(os.getenv("PGPORT", "5432")),
    "dbname": os.getenv("PGDATABASE", "gulfstream"),
    "user": os.getenv("PGUSER", "ohanafy"),
    "password": os.getenv("PGPASSWORD", "Xq7!mR#2vK$9pLw@nZ"),
    "sslmode": "require",
}

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "VIP_Pricing_Data_Export.xlsx")

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="4472C4")
CURRENCY_FMT = "$#,##0.00"
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


def format_vip_date(val):
    """Convert VIP numeric date (YYYYMMDD int) to readable string."""
    if val is None:
        return ""
    try:
        s = str(int(val))
        if len(s) == 8:
            return f"{s[4:6]}/{s[6:8]}/{s[0:4]}"
    except (ValueError, TypeError):
        pass
    return str(val)


def auto_width(ws, min_width=10, max_width=45):
    """Estimate and set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        width = max(min(max_len + 3, max_width), min_width)
        ws.column_dimensions[col_letter].width = width


def style_header(ws, num_cols):
    """Apply header styling to the first row of a worksheet."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_rows(ws, start_row=2, currency_cols=None, date_cols=None):
    """Apply alternating shading + formatting to data rows."""
    currency_cols = currency_cols or set()
    date_cols = date_cols or set()
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        row_num = row[0].row
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if (row_num - start_row) % 2 == 0:
                cell.fill = ALT_ROW_FILL
            col_idx = cell.column
            if col_idx in currency_cols and cell.value is not None:
                cell.number_format = CURRENCY_FMT
            if col_idx in date_cols and cell.value is not None:
                cell.value = format_vip_date(cell.value)


def write_sheet(ws, headers, rows, currency_cols=None, date_cols=None):
    """Write headers + data to a worksheet and apply formatting."""
    ws.append(headers)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(list(r))
    style_data_rows(ws, start_row=2, currency_cols=currency_cols, date_cols=date_cols)
    auto_width(ws)


def fetch(cur, sql):
    """Execute SQL and return all rows."""
    cur.execute(sql)
    return cur.fetchall()


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("Connecting to database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        print(f"ERROR connecting: {e}")
        sys.exit(1)
    cur = conn.cursor()
    print("Connected.\n")

    wb = Workbook()

    # -----------------------------------------------------------------------
    # 1. Summary sheet
    # -----------------------------------------------------------------------
    print("Building Summary sheet...")
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title row
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary.cell(row=1, column=1, value="VIP Pricing Data Export \u2014 Gulf Distributing")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws_summary.row_dimensions[1].height = 36

    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws_summary.cell(row=2, column=1).font = Font(italic=True, color="808080")
    ws_summary.cell(row=3, column=1, value="")

    # Pricing tables overview
    pricing_tables = [
        ("HDRPRCODT", "Price Codes", "Define price code categories for customers"),
        ("HDRPRGRPT", "Price Groups", "Group products into pricing tiers"),
        ("DISCWKSTT", "Discount Worksheets / Pricelists", "Master pricelist definitions with dates & types"),
        ("PENDPRICT", "Pending Prices / Pricelist Items", "Individual product pricing within pricelists"),
        ("DISCOUTT", "Discounts / Promotions", "Active and historical discount assignments"),
        ("DPMASTT", "Deployment Master (header)", "Price deployment headers by product/code/warehouse"),
        ("DPMAST1T", "Deployment Master (detail)", "Front-line prices, selling tiers, rebate amounts"),
        ("CCNAMET", "Cost Component Names", "Named cost layers (FOB, freight, tax, etc.)"),
        ("DEPOSITST", "Deposits & Fees", "Container deposits and miscellaneous fees"),
        ("COMMPROFT", "Commission Profiles", "Sales rep commission rate profiles"),
        ("MQTCOST", "Monthly/Quarterly Costs", "Periodic cost snapshots by warehouse/item"),
        ("FLPWKST", "Front-Line Pricing Worksheet", "Temporary front-line price overrides"),
    ]

    row_start = 4
    ws_summary.cell(row=row_start, column=1, value="Table").font = HEADER_FONT
    ws_summary.cell(row=row_start, column=1).fill = HEADER_FILL
    ws_summary.cell(row=row_start, column=2, value="Name").font = HEADER_FONT
    ws_summary.cell(row=row_start, column=2).fill = HEADER_FILL
    ws_summary.cell(row=row_start, column=3, value="Row Count").font = HEADER_FONT
    ws_summary.cell(row=row_start, column=3).fill = HEADER_FILL
    ws_summary.cell(row=row_start, column=4, value="Description").font = HEADER_FONT
    ws_summary.cell(row=row_start, column=4).fill = HEADER_FILL

    for i, (tbl, name, desc) in enumerate(pricing_tables):
        r = row_start + 1 + i
        try:
            cur.execute(f'SELECT COUNT(*) FROM staging."{tbl}" WHERE import_is_deleted = false')
            cnt = cur.fetchone()[0]
        except Exception:
            conn.rollback()
            cnt = "N/A"
        ws_summary.cell(row=r, column=1, value=tbl)
        ws_summary.cell(row=r, column=2, value=name)
        ws_summary.cell(row=r, column=3, value=cnt)
        ws_summary.cell(row=r, column=3).number_format = "#,##0"
        ws_summary.cell(row=r, column=4, value=desc)
        if i % 2 == 0:
            for c in range(1, 5):
                ws_summary.cell(row=r, column=c).fill = ALT_ROW_FILL

    # Key stats section
    stats_start = row_start + len(pricing_tables) + 3
    ws_summary.cell(row=stats_start, column=1, value="Key Statistics").font = SUBTITLE_FONT
    stat_queries = [
        (
            "Total Items with Pricing",
            'SELECT COUNT(DISTINCT TRIM("PRODID")) FROM staging."PENDPRICT" WHERE import_is_deleted = false',
        ),
        (
            "Total Pricelists (Disc Worksheets)",
            'SELECT COUNT(*) FROM staging."DISCWKSTT" WHERE import_is_deleted = false AND TRIM("DISCOUNTID") <> \'\'',
        ),
        (
            "Active Promotions (end >= today)",
            'SELECT COUNT(*) FROM staging."DISCOUTT" WHERE import_is_deleted = false AND "ENDDATE" >= 20260313',
        ),
        (
            "Distinct Price Codes",
            'SELECT COUNT(DISTINCT TRIM("PRICECODE")) FROM staging."HDRPRCODT" WHERE import_is_deleted = false',
        ),
        (
            "Distinct Price Groups",
            'SELECT COUNT(DISTINCT TRIM("PRICEGROUP")) FROM staging."HDRPRGRPT" WHERE import_is_deleted = false',
        ),
        ("Deployment Pricing Records", 'SELECT COUNT(*) FROM staging."DPMASTT" WHERE import_is_deleted = false'),
        (
            "Fee / Deposit Records (non-zero)",
            'SELECT COUNT(*) FROM staging."DEPOSITST" WHERE import_is_deleted = false AND "DPDEPO" IS NOT NULL AND "DPDEPO" <> 0',
        ),
        ("Commission Profiles", 'SELECT COUNT(*) FROM staging."COMMPROFT" WHERE import_is_deleted = false'),
    ]
    for j, (label, sql) in enumerate(stat_queries):
        r = stats_start + 1 + j
        try:
            cur.execute(sql)
            val = cur.fetchone()[0]
        except Exception:
            conn.rollback()
            val = "N/A"
        ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=r, column=2, value=val)
        if isinstance(val, (int, float)):
            ws_summary.cell(row=r, column=2).number_format = "#,##0"
        if j % 2 == 0:
            for c in range(1, 3):
                ws_summary.cell(row=r, column=c).fill = ALT_ROW_FILL

    auto_width(ws_summary)
    ws_summary.freeze_panes = "A5"
    print("  Summary done.")

    # -----------------------------------------------------------------------
    # 2. Price Codes
    # -----------------------------------------------------------------------
    print("Building Price Codes sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("PRICECODE"), TRIM("DESCRIPTION"), TRIM("PCCOSTNAME"),
               "ALLOWDISCOUNTS", "CHECKZEROPRICE"
        FROM staging."HDRPRCODT"
        WHERE import_is_deleted = false
        ORDER BY "PRICECODE"
    """,
    )
    ws = wb.create_sheet("Price Codes")
    write_sheet(ws, ["Price Code", "Description", "Cost Name", "Allow Discounts", "Check Zero Price"], rows)
    print(f"  Price Codes: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 3. Price Groups
    # -----------------------------------------------------------------------
    print("Building Price Groups sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("PRICEGROUP"), TRIM("DESCRIPTION"), TRIM("PGCOSTNAME"), "PDAFLAG"
        FROM staging."HDRPRGRPT"
        WHERE import_is_deleted = false
        ORDER BY "PRICEGROUP"
    """,
    )
    ws = wb.create_sheet("Price Groups")
    write_sheet(ws, ["Price Group", "Description", "Cost Name", "PDA Flag"], rows)
    print(f"  Price Groups: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 4. Pricelists (Disc Worksheets)
    # -----------------------------------------------------------------------
    print("Building Pricelists sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("DISCOUNTID"), TRIM("DISCDESC"), TRIM("DISCOUNTTYPE"),
               TRIM("DISCLVL"), TRIM("BYBOTTLE"), "STARTDATE", "ENDDATE",
               TRIM("INDIPRICE"), TRIM("PERFDISC")
        FROM staging."DISCWKSTT"
        WHERE import_is_deleted = false AND TRIM("DISCOUNTID") <> ''
        ORDER BY "STARTDATE" DESC
    """,
    )
    ws = wb.create_sheet("Pricelists (Disc Worksheets)")
    write_sheet(
        ws,
        [
            "Discount ID",
            "Description",
            "Type",
            "Level",
            "By Bottle",
            "Start Date",
            "End Date",
            "Individual Price",
            "Performance Discount",
        ],
        rows,
        date_cols={6, 7},
    )
    print(f"  Pricelists: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 5. Pricelist Items (Pending)
    # -----------------------------------------------------------------------
    print("Building Pricelist Items sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("PRODID"), TRIM("CODEID"), TRIM("DISCCODE"),
               "CASEPRICE", "EACHPRICE", "POSTOFF",
               TRIM("DISCGRP"), "STARTDATE", "ENDDATE", TRIM("TYPEID")
        FROM staging."PENDPRICT"
        WHERE import_is_deleted = false
        ORDER BY TRIM("PRODID"), TRIM("CODEID")
    """,
    )
    ws = wb.create_sheet("Pricelist Items (Pending)")
    write_sheet(
        ws,
        [
            "Product ID",
            "Price Code",
            "Discount Code",
            "Case Price",
            "Each Price",
            "Post-Off",
            "Discount Group",
            "Start Date",
            "End Date",
            "Type",
        ],
        rows,
        currency_cols={4, 5, 6},
        date_cols={8, 9},
    )
    print(f"  Pricelist Items: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 6. Promotions (Active)
    # -----------------------------------------------------------------------
    print("Building Promotions (Active) sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("PRODID"), TRIM("CODEID"), TRIM("DISCCODE"),
               TRIM("TYPEID"), TRIM("DISCGRP2"), TRIM("DISCGRP3"), TRIM("DISCGRP4"),
               "STARTDATE", "ENDDATE"
        FROM staging."DISCOUTT"
        WHERE import_is_deleted = false AND "ENDDATE" >= 20260313
        ORDER BY TRIM("DISCCODE"), TRIM("PRODID")
        LIMIT 10000
    """,
    )
    ws = wb.create_sheet("Promotions (Active)")
    write_sheet(
        ws,
        [
            "Product ID",
            "Price Code",
            "Discount Code",
            "Type",
            "Disc Group 2",
            "Disc Group 3",
            "Disc Group 4",
            "Start Date",
            "End Date",
        ],
        rows,
        date_cols={8, 9},
    )
    print(f"  Promotions (Active): {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 7. Deployment Master (Sample)
    # -----------------------------------------------------------------------
    print("Building Deployment Master sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM(d."DPITEM"), TRIM(d."DPPRCD"), TRIM(d."DPPGRP"),
               TRIM(d."DPWHSE"), TRIM(d."DPMETHOD"),
               d."DPBDAT", d."DPEDAT",
               p."FRONTLINEPRICE", p."MINIMUMREBATEPRICE", p."REBATEPERCENT",
               p."SELLINGPRICE01", p."SELLINGPRICE02", p."SELLINGPRICE03",
               p."REBATEAMOUNT01", p."REBATEAMOUNT02", p."REBATEAMOUNT03"
        FROM staging."DPMASTT" d
        JOIN staging."DPMAST1T" p ON d."DPIDENTITY" = p."ID_DPMAST"
        WHERE d.import_is_deleted = false AND p.import_is_deleted = false
        ORDER BY d."DPITEM", d."DPPRCD"
        LIMIT 500
    """,
    )
    ws = wb.create_sheet("Deployment Master (Sample)")
    write_sheet(
        ws,
        [
            "Product ID",
            "Price Code/Group",
            "Type",
            "Warehouse",
            "Method",
            "Start Date",
            "End Date",
            "Front Line Price",
            "Min Rebate Price",
            "Rebate %",
            "Sell Price 01",
            "Sell Price 02",
            "Sell Price 03",
            "Rebate Amt 01",
            "Rebate Amt 02",
            "Rebate Amt 03",
        ],
        rows,
        currency_cols={8, 9, 11, 12, 13, 14, 15, 16},
        date_cols={6, 7},
    )
    print(f"  Deployment Master: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 8. Cost Components
    # -----------------------------------------------------------------------
    print("Building Cost Components sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("CNCOST"), TRIM("CNNAME")
        FROM staging."CCNAMET"
        WHERE import_is_deleted = false
        ORDER BY "CNCOST"
    """,
    )
    ws = wb.create_sheet("Cost Components")
    write_sheet(ws, ["Cost Code", "Cost Name"], rows)
    print(f"  Cost Components: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 9. Deposits & Fees
    # -----------------------------------------------------------------------
    print("Building Deposits & Fees sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("DPDEPTYP"), TRIM("DPITEM"), TRIM("DPACCT"),
               "DPDEPO", "DPSDAT", "DPEDAT",
               TRIM("DPDCOD"), TRIM("DPTYPE")
        FROM staging."DEPOSITST"
        WHERE import_is_deleted = false
          AND "DPDEPO" IS NOT NULL AND "DPDEPO" <> 0
        ORDER BY TRIM("DPDEPTYP"), "DPDEPO" DESC
        LIMIT 2000
    """,
    )
    ws = wb.create_sheet("Deposits & Fees")
    write_sheet(
        ws,
        [
            "Deposit Type",
            "Item Code",
            "Account/Price Code",
            "Deposit Amount",
            "Start Date",
            "End Date",
            "Deposit Code",
            "Type",
        ],
        rows,
        currency_cols={4},
        date_cols={5, 6},
    )
    print(f"  Deposits & Fees: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 10. Commissions
    # -----------------------------------------------------------------------
    print("Building Commissions sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("CPID"), TRIM("CPDESC"), "CPRATE",
               TRIM("CPTYPE"), "CPBAMT", "CPMINPAY",
               TRIM("CPQPAY"), TRIM("CPPSET")
        FROM staging."COMMPROFT"
        WHERE import_is_deleted = false
        ORDER BY "CPID"
    """,
    )
    ws = wb.create_sheet("Commissions")
    write_sheet(
        ws,
        ["Profile ID", "Description", "Rate", "Type", "Base Amount", "Min Pay", "Qty Pay", "Product Set"],
        rows,
        currency_cols={3, 5, 6},
    )
    print(f"  Commissions: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 11. Monthly Costs (Summary)
    # -----------------------------------------------------------------------
    print("Building Monthly Costs sheet...")
    rows = fetch(
        cur,
        """
        SELECT TRIM("WHSE"),
               COUNT(DISTINCT "ITEM"),
               ROUND(AVG("COST")::numeric, 2),
               ROUND(MIN("COST")::numeric, 2),
               ROUND(MAX("COST")::numeric, 2),
               COUNT(*)
        FROM staging."MQTCOST"
        WHERE import_is_deleted = false
        GROUP BY TRIM("WHSE")
        ORDER BY COUNT(*) DESC
    """,
    )
    ws = wb.create_sheet("Monthly Costs (Summary)")
    write_sheet(
        ws,
        ["Warehouse", "Item Count", "Avg Cost", "Min Cost", "Max Cost", "Total Records"],
        rows,
        currency_cols={3, 4, 5},
    )
    print(f"  Monthly Costs: {len(rows)} rows")

    # -----------------------------------------------------------------------
    # 12. VIP to Ohanafy Mapping
    # -----------------------------------------------------------------------
    print("Building VIP to Ohanafy Mapping sheet...")
    ws = wb.create_sheet("VIP to Ohanafy Mapping")

    mapping_data = [
        ("DISCWKSTT", "DISCOUNTID", "Pricelist__c", "Key__c", "TRIM"),
        ("DISCWKSTT", "DISCDESC", "Pricelist__c", "Name", "TRIM"),
        ("DISCWKSTT", "DISCOUNTTYPE", "Pricelist__c", "Type__c", "A\u2192Discount, F/I\u2192Individual Price"),
        ("DISCWKSTT", "DISCOUNTTYPE", "Pricelist__c", "Discount_Type__c", "P/%\u2192Percent, else\u2192Dollars"),
        ("DISCWKSTT", "STARTDATE", "Pricelist__c", "Start_Date__c", "YYYYMMDD\u2192Date"),
        ("DISCWKSTT", "ENDDATE", "Pricelist__c", "End_Date__c", "YYYYMMDD\u2192Date"),
        ("PENDPRICT", "PRODID", "Pricelist_Item__c", "Item__c", "Lookup by item code"),
        ("PENDPRICT", "DISCCODE", "Pricelist_Item__c", "Pricelist__c", "Lookup by Key__c"),
        ("PENDPRICT", "CASEPRICE", "Pricelist_Item__c", "Case_Price__c", "Direct"),
        ("PENDPRICT", "EACHPRICE", "Pricelist_Item__c", "Unit_Price__c", "Direct"),
        ("PENDPRICT", "POSTOFF", "Pricelist_Item__c", "Discount_Dollars__c", "Direct"),
        ("DISCOUTT", "DISCCODE", "Promotion__c", "Name", "Lookup description from DISCWKSTT"),
        ("DISCOUTT", "TYPEID", "Promotion__c", "(logic)", "G\u2192chain-level, C/A\u2192customer-specific"),
        ("DISCOUTT", "STARTDATE", "Promotion__c", "Start_Date__c", "YYYYMMDD\u2192Date"),
        ("DISCOUTT", "ENDDATE", "Promotion__c", "End_Date__c", "YYYYMMDD\u2192Date"),
        ("DISCOUTT", "PRODID", "Promotion_Item__c", "Item__c", "Lookup"),
        ("DEPOSITST", "DPDEPTYP", "Fee__c", "Type__c", "K\u2192Storage, S\u2192Misc. Charge"),
        ("DEPOSITST", "DPDEPO", "Fee__c", "Default_Amount__c", "Direct"),
        (
            "DPMAST1T",
            "FRONTLINEPRICE",
            "Item__c / Pricelist_Item__c",
            "Default_Case_Price__c / Case_Price__c",
            "Front-line price; may map to item default or pricelist item",
        ),
        (
            "DPMAST1T",
            "SELLINGPRICE01-20",
            "Pricelist_Item__c",
            "Case_Price__c",
            "Select active tier for each pricelist",
        ),
        ("DPMAST1T", "REBATEAMOUNT01-20", "Promotion__c", "Discount_Dollars__c", "Rebate as promotion discount"),
        (
            "HDRPRCODT",
            "(table)",
            "Pricelist_Account__c",
            "(assignment logic)",
            "Customer price code \u2192 pricelist assignment",
        ),
        ("HDRPRGRPT", "(table)", "Pricelist_Group__c", "(tier settings)", "Red Bull tiers \u2192 tier settings"),
        ("FLPWKST", "CASEPRICE", "Promotion__c", "Discounted_Case_Price__c", "Is_Front_Line=true"),
        ("COMMPROFT", "(table)", "(no direct mapping)", "-", "Reference only"),
        ("CCNAMET / CCSTANT", "(table)", "(no direct mapping)", "-", "Reference only"),
    ]

    write_sheet(ws, ["VIP Table", "VIP Field", "Ohanafy Object", "Ohanafy Field", "Transform Notes"], mapping_data)
    print(f"  VIP to Ohanafy Mapping: {len(mapping_data)} rows")

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    print(f"\nSaving workbook to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    file_size = os.path.getsize(OUTPUT_FILE)
    print(f"Done! File size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")
    print(f"Sheets: {wb.sheetnames}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
